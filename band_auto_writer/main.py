from ongo import *
import ctypes
from openpyxl import load_workbook

# global onth
# onth = threading.Thread(target=lambda: goScript(ipchk_var.get(), id_input.get()))
def th():
    getDict = {'ipval': ipVal.get(),}
    getDict['nlist'] = idbox.current() + 1
    onth = threading.Thread(target=lambda: goScript(getDict))
    onth.daemon = True
    onth.start()


# 윈도우 창 생성 및 버튼 화면 조절
root = Tk()
root.title("밴드 자동화")
root.geometry("240x210+700+300")
root.resizable(False, FALSE)

frame1 = LabelFrame(root, text='아이피 변경', padx=60, pady=5)  # padx / pady 내부여백
frame1.pack(padx=10, pady=5)  # padx / pady 외부여백

ipVal = IntVar()
ipChk1 = Radiobutton(frame1, text="아이피 변경", value=1, variable=ipVal)
ipChk2 = Radiobutton(frame1, text="아이피 미변경", value=0, variable=ipVal)
ipChk2.select()
ipChk1.pack()
ipChk2.pack()


frame4 = LabelFrame(root, text='아이디 선택', padx=30, pady=5)
frame4.pack(padx=10, pady=5)


# 엑셀파일 아이디 내열 list에 담기

wb = load_workbook('./content/etc/nid.xlsx')
ex = wb.active


nid_list = []
nlogin_list = []
i = 0
while True:
    i = i + 1
    id_val = ex.cell(i, 2).value
    if id_val is None:
        break
    else:
        nid_list.append(id_val)

idbox = ttk.Combobox(frame4, values=nid_list)
idbox.current(0)
idbox.pack()

# 시작 버튼 생성
btn1 = Button(root, text='시작하기', command=th, padx=50)
btn1.pack()

btn2 = Button(root, text="종료하기", command=exitApp, padx=50)
btn2.pack()

# ********************************


# 윈도우창 계속 띄우기
root.mainloop()