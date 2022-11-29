import linecache
import random
import threading
import time
import sys
import os

import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob


def goScript(getval):

    now_path = os.getcwd()

    # 텍스트파일 리스트 저장
    txt_list = []
    txt_file_pattern = rf'{now_path}\content\*.txt'
    all_txt_file = glob.glob(txt_file_pattern)
    for txt_file in all_txt_file:
        txt_name = txt_file.split('\\')[-1]
        txt_list.append(txt_name)

    # 이미지 파일 랜덤 저장
    img_list = []

    img_extension_list = ['jpg', 'jpeg', 'png']
    for img_ex in img_extension_list:
        img_file_pattern = rf'{now_path}\content\*.{img_ex}'
        all_img_file = glob.glob(img_file_pattern)
        for img_file in all_img_file:
            img_name = img_file.split('\\')[-1]
            img_list.append(img_name)

    # 아이피 변경 / 밴드 로그인

    if getval['ipval'] == 0:
        getIp = ''
    else:
        getIp = changeIp()

    with open('./content/etc/useragent/useragent_all.txt') as uatxt:
        ualine = len(uatxt.readlines())

    el = getval['nlist']
    wb = load_workbook('./content/etc/nid.xlsx')
    ex = wb.active

    if ex.cell(el, 1).value is None:
        ua_ranVal = random.randrange(1, ualine)
        ex.cell(el, 1).value = ua_ranVal
        wb.save('./content/etc/nid.xlsx')
        pg.alert(text='UserAgent 값이 랜덤으로 설정 되었습니다.')

    ua_data = linecache.getline('./content/etc/useragent/useragent_all.txt', ex.cell(el, 1).value).strip()

    options = Options()
    user_agent = ua_data
    options.add_argument('user-agent=' + user_agent)
    global driver



    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(chrome_options=options, service=service)

    driver.get('https://auth.band.us/login_page')

    nlogin_btn = searchElement('.-naver')

    nlogin_btn[0].click()
    wait_float(3, 5)

    searchElement("#id")

    nid = ex.cell(el, 2).value
    npwd = ex.cell(el, 3).value

    focus_window('로그인')
    bandNaverLogin(nid, npwd)
    
    cookie_btn = searchElement('.bcm_btnCookieOn')
    cookie_btn[0].click()
    wait_float(2, 4)

    wb = load_workbook("./content/etc/band_list.xlsx")
    ws = wb.worksheets[0]

    while True:
        for i in range(1, ws.max_row + 1):
            driver.get(ws.cell(i, 1).value)
            print('창 이동 완료!! 다음 밴드 작성 준비')
            searchElement('.groupBandInfoBox')
            
            # joinBand(ex, el)
            # 마우스 이동
            
            allRand = random.randrange(2,5)
            for i in range(1,allRand):
                pg.moveTo(350, 550)
                wait_float(0.7,1.5)
                onRand = random.randrange(2,5)
                for k in range(1, onRand):
                    pg.scroll(-1000)
                    wait_float(0.7,1.5)
                pg.click()
                wait_float(1.2,2.5)

                
                try:
                    closeBtn = driver.find_elements(by=By.CSS_SELECTOR, value='.btnViewerClose')
                    if closeBtn:
                        untilEleGone('.btnViewerClose', '.btnViewerClose')
                except:
                    pass
                
                try:
                    closeBtn = driver.find_elements(by=By.CSS_SELECTOR, value='._postDetailListKeepingRegion.-show')
                    if closeBtn:
                        untilEleGone('.aIconClose', '._postDetailListKeepingRegion.-show')
                except:
                    pass

                
                print('전체 문제 없음~~~')
            

            if len(driver.window_handles) > 1:
                wait_float(0.3,0.8)
                driver.switch_to.window(driver.window_handles[1])
                wait_float(0.3,0.8)
                driver.close()
                wait_float(0.3,0.8)
                driver.switch_to.window(driver.window_handles[0])
            print('창이 문제인건가???')
            

            
            print('버튼 찾기 시작~~~')

            # write_btn = searchElement('._writePostButton')
            # print('버튼은 어케됨??')
            
            # if not write_btn:
            #     driver.close()
            #     first_tab = driver.window_handles[0]
            #     driver.switch_to.window(window_name=first_tab)
            #     continue
            # write_btn[0].click()

            while True:
                try:
                    writeBtn = driver.find_element(by=By.CSS_SELECTOR, value='._writePostButton')
                    writeBtn.click()
                except:
                    pass
                try:
                    writeBtn = driver.find_element(by=By.CSS_SELECTOR, value='._btnPostWrite')
                    writeBtn.click()
                except:
                    pass
                wait_float(1.1, 1.7)

                try:
                    txt_area = driver.find_element(by=By.CSS_SELECTOR, value='._postTextArea')
                    txt_area.click()
                    break
                except:
                    continue
            
            print('버튼 찾기 끝~~~')

            # 글쓰기 시작

            focus_window('밴드')
            if txt_list:
                txt_ran = random.randrange(0, len(txt_list))
                ran_txt_file = txt_list[txt_ran]
                with open(f"{now_path}\content\{ran_txt_file}", 'r') as txtfile:
                    pyperclip.copy(txtfile.read())
                    wait_float(0.5, 1.0)
                    pg.hotkey('ctrl', 'v')
                    # for line in txtfile.readlines():
                    #     keyboard.write(line, 0.2)

            if img_list:
                wait_float(1.5, 2.7)
                img_ran = random.randrange(0, len(img_list))
                ran_img_file = img_list[img_ran]
                img_upload_btn = driver.find_elements(by=By.CSS_SELECTOR, value='._photoUploadDiv')
                img_upload_btn[0].click()
                wait_float(1.5, 2.7)

                keyboard.write(f"{now_path}\content")
                wait_float(1, 2)
                pg.press('enter')
                wait_float(1, 2)
                keyboard.write(ran_img_file)
                wait_float(1, 2)
                pg.press('enter')
                wait_float(3, 4)
            
            try:
                btnEle = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '._btnPostWrite.-active')))
                btnEle.click()
                wait_float(3, 4)
            except:
                pass


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell("am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell("am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(5)
        getIp = requests.get("http://ip.jsontest.com").json()['ip']
    except:
        time.sleep(5)
        getIp = requests.get("http://ip.jsontest.com").json()['ip']
    return getIp


def searchElement(ele):
    time.sleep(1)
    re_count = 1
    element = ""
    while True:
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            re_count += 1

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.5, 1.2)
    return selected_element


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winname: str):
    win_list = gw.getAllTitles()
    get_list = [list for list in win_list if winname in list]
    win = gw.getWindowsWithTitle(get_list[0])[0]
    win.activate()

def bandNaverLogin(nid, npwd):
    while True:
        pyperclip.copy(nid)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'a')
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'v')
        wait_float(0.5, 1.0)

        pyperclip.copy(npwd)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'a')
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'v')
        wait_float(0.5, 1.0)

        id_input_value = id_input[0].get_attribute('value')
        if id_input_value == nid:
            break

    submit_btn = driver.find_elements(by=By.CSS_SELECTOR, value='.btn_check')
    submit_btn[0].click()
    wait_float(3, 5)

    pg.alert(title='인증 대기', text='휴대폰 인증을 완료 해주세요!')
    

def joinBand(ex, el):
    chk_join_band = driver.find_elements(by=By.CSS_SELECTOR, value='._btnBandJoin')
    if chk_join_band:
        chk_join_band[0].click()
        wait_float(1, 2)
        continue_browser = driver.find_element(by=By.CSS_SELECTOR, value='._btnContinueInBrowser')
        continue_browser.click()
        wait_float(1, 2)

        joinAnswer = driver.find_elements(by=By.CSS_SELECTOR, value='._joinAnswer')
        if joinAnswer:
            joinAnswer[0].send_keys('가입합니다.')
            wait_float(1, 2)
            btnConfirm = driver.find_element(by=By.CSS_SELECTOR, value='._btnConfirm')
            btnConfirm.click()
            wait_float(1, 2)

        profile_list = driver.find_elements(by=By.CSS_SELECTOR, value='.profileBox')
        if not profile_list:
            driver.close()
            first_tab = driver.window_handles[0]
            driver.switch_to.window(window_name=first_tab)
            return 'continue'

        find_profile = ""
        for profile in profile_list:
            if profile.text == ex.cell(el, 4).value:
                find_profile = "ok"
                profile.click()
                wait_float(1, 2)

        if find_profile == "":
            newprofile = driver.find_elements(by=By.CSS_SELECTOR, value='._btnNewProfile')
            newprofile[0].click()
            wait_float(1, 2)
            profile_name = driver.find_elements(by=By.CSS_SELECTOR, value='._nameInput')
            profile_name[0].send_keys(ex.cell(el, 4).value)
            wait_float(1, 2)

        join_confirmbtn = driver.find_elements(by=By.CSS_SELECTOR, value='._btnConfirm')
        join_confirmbtn[0].click()
        wait_float(1, 2)
        join_alert = driver.switch_to.alert
        join_alert.accept()
        wait_float(3, 4)

def untilEleShow(clickEle,searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
            try:
                btnEle = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
                if btnEle is not None:
                    return
            except:
                continue
        except:
            pass
        
def untilEleGone(clickEle,searchEle):
    while True:
        print('돌고 있음 문제되는 부분은??' + clickEle)
        closeBtn = driver.find_element(by=By.CSS_SELECTOR, value=clickEle)
        try:
            closeBtn = driver.find_element(by=By.CSS_SELECTOR, value=clickEle)
            closeBtn.click()
            wait_float(0.8,1.9)
        except:
            pass
        

        try:
            driver.find_element(by=By.CSS_SELECTOR, value=searchEle)
            wait_float(0.5,0.9)
            continue
        except:
            return
        # try:
        #     btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, clickEle)))
        #     btnEle.click()
        #     time.sleep(1)
        # except:
        #     pass
        
        # try:
        #     btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
        #     if btnEle is None:
        #         return
        #     else:
        #         continue
        # except:
        #     return