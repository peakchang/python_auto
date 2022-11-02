import random
import threading
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
from typing import Optional
from pyparsing import And
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import aiohttp
import asyncio


# chrome://version/ 에서 '프로필경로' 복사, 난 왜 디폴트만 되지?? 뭔... 딴건 필요 없쓰....


def goScript(getDict):
    
    if getDict['nlist'] == 1:
        pg.alert('아이디가 선택되지 않았습니다. 다시 실행해주세요')
        exitApp()
    
    global driver
    
    
    
    
    
    
    files = sorted(glob.glob('.\\etc\\content\\*.txt'), key=os.path.getctime)
    for file in files:
        print(file)
        setFileName = file.replace('.\\etc\\content\\', '')
        print(setFileName)
        with open(file, 'r') as f:
            f.read()
            # print(f.read())
    
    
    exLineNum = getDict['nlist']
    wb = load_workbook('./etc/nid.xlsx')
    ex = wb.active
    
    
    preIp = ''
    
    if getDict['ipval'] == 1:
        while True:
            getIP = changeIp()
            print(getIP)
            if not preIp == getIP:
                preIp = getIP
                break
        
    
    
    
    options = Options()
    user_data = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    service = Service(ChromeDriverManager().install())
    options.add_argument(f"user-data-dir={user_data}")
    # options.add_argument('--profile-directory=Profile 3')
    pg.alert('프로필을 선택해주세요! 그전에 글 / 댓글 바꿨는지 꼭 확인 하세요! 쫌 해!')
    driver = webdriver.Chrome(service=service, chrome_options=options)
    
    driver.get('https://www.naver.com')
    
    
    
    loginBtn = searchElement('.sc_login')
    loginBtn[0].click()
    
    # cb.copy(ex.cell(exLineNum, 1).value)
    # pg.alert('아이디가 복사되었습니다. 붙여넣기 해주세요')
    
    # cb.copy(ex.cell(exLineNum, 2).value)
    # pg.alert('비밀번호가 복사되었습니다. 로그인 후 블로그 작성 준비가 되면 엔터를 클릭해주세요!')
    
    # while True:
    
    searchElement('#id')
    focus_window('chrome')
    wait_float(0.3,0.9)
    while True:
        
        pg.click(400,500)
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        inputId.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 1).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        if inputId.get_attribute('value') != "":
            break
        
    while True:
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        inputPw.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 2).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        if inputPw.get_attribute('value') != "":
            break
    
    btnLogin = searchElement('.btn_login')
    btnLogin[0].click()
    
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
    
    navItem = searchElement('.nav_item')
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item')
    menu_my_blog[1].click()
    
    driver.switch_to.window(driver.window_handles[1])
    
    
    path_dir = './etc/content'
    # folder_list = os.listdir(path_dir)
    # for folder in folder_list:
    
    
    
    files = sorted(glob.glob('.\\etc\\content\\*.txt'), key=os.path.getctime)
    writeCount = 0
    for file in files:
        writeCount += 1
        pg.alert(f'{writeCount}번째 글쓰기를 시작합니다!!')
        # driver.to_switch()
        driver.switch_to.window(driver.window_handles[1])
        
        driver.switch_to.frame('mainFrame')
        
        writeArea = searchElement('.se-component-content')
        
        with open(file, 'r') as f:
            getLines = f.readlines()
            
        for i, getline in enumerate(getLines):
            focus_window('블로그')
            getline = getline.replace('\n', '')
            chkImg = getline.split('|')
            if chkImg[0] == 'img_line':
                nowPath = os.getcwd()
                
                img_btn = searchElement('.se-image-toolbar-button')
                img_btn[0].click()
                wait_float(1.5,2.3)
                
                imagePath = nowPath + f"\etc\content"
                wait_float(1.5, 2.2)
                pyperclip.copy(imagePath)
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                
                wait_float(0.9, 1.6)
                pyperclip.copy(chkImg[1])
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(3.5,4.5)
                continue
            
            if i == 0:
                writeArea[0].click()
                keyboard.write(text=getline, delay=0.05)
                wait_float(1.2,2.8)
            elif i == 1:
                writeArea[1].click()
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
            else:
                keyboard.write(text=getline, delay=0.05)
                wait_float(0.5,0.9)
                pg.press('enter')
                wait_float(0.5,0.9)
        pg.alert('글 작성 완료!! 글쓰기 완료 버튼 클릭 후 확인을 눌러주세요!')
        
    
    chkVal = pg.confirm(text='글쓰기가 완료 되었습니다!! 댓글을 진행 하시겠습니까?', buttons=['go','stop'])
    if chkVal == 'go':
        getUrl = searchElement('._transPosition')
        getUrl[0].click()
        wait_float(1.5,2.5)
        pg.press('enter')
        driver.switch_to.default_content()
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        goToNaverMain = searchElement('.link_naver')
        goToNaverMain[0].click()
        pg.alert('일단 블로그 링크는 URL 복사 클릭 해주시고 네이버 메인에서 대기하세요!')
        blogReplyWork()
    else:
        exitApp()
        
    
# def makeBlogContent():
    
#     with open('./etc/blog_link.txt', 'r') as f:
#         getBlogLink = f.readlines()
    
#     try:
#         subject = getBlogLink[2].replace('\n', '')
#         with open('./etc/text.txt', 'w') as f:
#             f.write(f'{subject}\n')
#     except:
#         pass
    
#     firstLinkTemp = getBlogLink[0].replace('\n', '')
#     firstLink = firstLinkTemp.replace('//','//m.')
#     firstPage = requests.get(firstLink)
#     fisrtContent = makeContentArr(firstPage)

#     secondLinkTemp = getBlogLink[0].replace('\n', '')
#     secondLink = secondLinkTemp.replace('//','//m.')
#     secondPage = requests.get(secondLink)
#     secondContent = makeContentArr(secondPage)
    
#     if len(fisrtContent) > len(secondContent):
#         contentLength = len(secondContent)
#     else:
#         contentLength = len(fisrtContent)
        
#     with open('./etc/text.txt', 'a') as f:
#         for i in range(0, contentLength):
#             getLastSentence_first = fisrtContent[i].pop(len(fisrtContent[i]) - 1)
#             getLastSentence_second = secondContent[i].pop(len(secondContent[i]) - 1)

            
#             if random.randrange(1,3) == 1:
#                 lastSentence = getLastSentence_first
#                 allArr = secondContent[i] + fisrtContent[i]
#             else:
#                 lastSentence = getLastSentence_second
#                 allArr = fisrtContent[i] + secondContent[i]
                
            
#             sampleRandom = random.sample(range(0, len(allArr)), len(allArr) // 3 * 2)
#             resultSentence = ''
            
#             for k in sampleRandom:
#                 resultSentence = resultSentence + allArr[k] + ' '
#             # for k in range(1, len(sampleRandom)):
#             #     resultSentence = resultSentence + allArr[sampleRandom[]] + ' '
#             resultSentence = resultSentence + lastSentence + '.'
            
#             f.write(f'{resultSentence}\n')
#     exitApp()







def blogReplyReady(getValList):
    
    if getValList['nlist'] == 1:
        pg.alert('아이디가 선택되지 않았습니다. 다시 실행해주세요')
        exitApp()
    
    global driver
    
    exLineNum = getValList['nlist']
    wb = load_workbook('./etc/nid.xlsx')
    ex = wb.active
    
    options = Options()
    user_data = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    service = Service(ChromeDriverManager().install())
    options.add_argument(f"user-data-dir={user_data}")
    # options.add_argument('--profile-directory=Profile 3')
    pg.alert('프로필을 선택해주세요!')
    driver = webdriver.Chrome(service=service, chrome_options=options)
    
    driver.get('https://www.naver.com')
    
    loginBtn = searchElement('.sc_login')
    loginBtn[0].click()
    
    # while True:
    
    searchElement('#id')
    focus_window('chrome')
    wait_float(0.3,0.9)
    while True:
        
        pg.click(400,500)
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        inputId.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 1).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
        if inputId.get_attribute('value') != "":
            break
        
    while True:
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        inputPw.click()
        wait_float(0.3,0.9)
        cb.copy(ex.cell(exLineNum, 2).value)
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'a')
        wait_float(0.3,0.9)
        pg.hotkey('ctrl', 'v')
        inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
        if inputPw.get_attribute('value') != "":
            break
    
    btnLogin = searchElement('.btn_login')
    btnLogin[0].click()
    
    pg.alert('대기요~~~~~')
    # 블로그 링크따기
    
    while True:
        try:
            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
            break
        except:
            driver.get('https://www.naver.com')
    
    navItem = searchElement('.nav_item')
    for mitem in navItem:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    menu_my_blog = searchElement('.menu_my_blog .item')
    menu_my_blog[0].click()
    
    driver.switch_to.window(driver.window_handles[1])
    driver.switch_to.frame('mainFrame')
    
    getUrl = searchElement('._transPosition')
    getUrl[0].click()
    wait_float(1.5,2.5)
    pg.press('enter')
    driver.switch_to.default_content()
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    
    goToNaverMain = searchElement('.link_naver')
    goToNaverMain[0].click()
    
    # 블로그 링크 따기 끝~~~
    
    blogReplyWork()




def blogReplyWork():
    navItem = searchElement('.nav_item')
    for mitem in navItem:
        if mitem.text == '카페':
            mitem.click()
            break
    
    cafeList = searchElement('.user_mycafe_info')
    getInCafe = ""
    for cafeOn in cafeList:
        if "소셜공간" in cafeOn.text:
            cafeOn.click()
            getInCafe = "on"
            break
    
    if getInCafe == "":
        driver.get('https://cafe.naver.com/sens3')
    else:
        driver.switch_to.window(driver.window_handles[0])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
    cafeWriteBtn = searchElement('.cafe-write-btn')
    if "가입" in cafeWriteBtn[0].text:
        pg.alert('카페 가입하기~~')
    
    workBoardLink = searchElement('#menuLink226')
    workBoardLink[0].click()
    
    driver.switch_to.frame('cafe_main')
    
    
    
    # 카페에 글 작성하기
    
    workBoardWriteBtn = searchElement('#writeFormBtn')
    workBoardWriteBtn[0].click()
    
    wait_float(1.5,2.5)
    driver.switch_to.window(driver.window_handles[1])
    
    
    subjectArea = searchElement('.FlexableTextArea')
    subjectArea[0].click()
    
    with open('./etc/social_cafe_content.txt', 'r') as r:
        cafeContent = r.readlines()
    
    keyboard.write(text=cafeContent[0], delay=0.05)
    wait_float(0.3,0.9)
    pg.moveTo(750,850)
    wait_float(0.3,0.9)
    pg.click()
    wait_float(0.3,0.9)
    pg.hotkey('ctrl', 'a')
    
    for i,conLine in enumerate(cafeContent):
        if i == 0:
            continue
        keyboard.write(text=conLine, delay=0.03)
        wait_float(0.5,1.5)
    pg.press('enter')
    pg.hotkey('ctrl', 'v')
    wait_float(1.5,2.5)
    BaseButton = searchElement('.BaseButton')
    BaseButton[0].click()
    
    
    
    
    wait_float(1.5,2.5)
    driver.switch_to.frame('cafe_main')
    
    buttonUrl = searchElement('.button_url')
    buttonUrl[0].click()
    
    wait_float(0.5,0.9)
    driver.close()
    
    driver.switch_to.window(driver.window_handles[0])
    
    
    
    workCafeLink = pyperclip.paste()
    workCafeNum = workCafeLink.split('/')[-1]
    preNick = ""
    
    # 카페에 글 작성하기 끝~
    
    
    
    
    forVal = random.randrange(6,8)
    
    ici = 0
    while True:
        ici += 1
        driver.switch_to.default_content()
        
        wait_float(0.3,0.9)
        workBoardLink = searchElement('#menuLink226')
        workBoardLink[0].click()
        
        wait_float(0.3,0.9)
        
        driver.switch_to.frame('cafe_main')
        
        wait_float(0.3,0.9)
        
        articleDiff = searchElement('.article-board')
        articleList = articleDiff[1].find_elements(by=By.CSS_SELECTOR, value=".td_article")
        
        
        
        if str(workCafeNum) in articleList[ici].find_element(by=By.CSS_SELECTOR, value=".board-number").text:
            forVal = forVal + 1
            continue
        
        wait_float(0.3,0.9)
        clickArticleTarget = articleList[ici].find_element(by=By.CSS_SELECTOR, value=".article")
        wait_float(0.3,0.9)
        # clickArticleTarget.click()
        untilEleShow(clickArticleTarget, '.nickname')
        
        
        nickname = searchElement('.nickname')
        if preNick == nickname[0].text:
            driver.back()
            wait_float(2.1,3.7)
            continue
        else:
            preNick = nickname[0].text
        
        chkLinkTag = driver.find_elements("xpath", "//*[contains(@class, 'se-fs-')]")

        for chkLink in chkLinkTag:
            try:
                getOtherBlogLink = chkLink.find_element(by=By.CSS_SELECTOR, value="a").get_attribute('href')
                if 'blog' in str(getOtherBlogLink):
                    chkOtherBlogLink = getOtherBlogLink.split('/')
                    if len(chkOtherBlogLink) < 5:
                        forVal = forVal + 1
                        driver.back()
                        wait_float(2.1,3.7)
                        break
                    else:
                        
                        chkLink.click()
                        
                        if 'm.' in getOtherBlogLink:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            gongamBtn = searchElement('.u_ico')
                            wait_float(0.3,0.9)
                            gongamBtn[-1].click()
                            wait_float(0.3,0.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                            
                        else:
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[1])
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.switch_to.frame('mainFrame')
                            wait_float(0.3,0.9)
                            gongamBtn = searchElement('.u_ico')
                            gongamBtn[1].click()
                            wait_float(0.5,1.5)
                            driver.switch_to.default_content()
                            wait_float(0.3,0.9)
                            driver.close()
                            wait_float(0.3,0.9)
                            driver.switch_to.window(driver.window_handles[0])
                            wait_float(0.3,0.9)
                        
                        driver.switch_to.frame('cafe_main')
                        replyArea = searchElement('.comment_inbox_text')
                        replyArea[0].click()
                        replyArea[0].send_keys('^0^')
                        pg.press('space')
                        
                        for i,conLine in enumerate(cafeContent):
                            if i == 0:
                                continue
                            keyboard.write(text=conLine, delay=0.03)
                            wait_float(0.5,1.5)
                        pg.press('enter')
                        pg.hotkey('ctrl', 'v')
                        wait_float(1.5,2.5)
                        
                        replySuccessBtn = searchElement('.btn_register')
                        driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
                        replySuccessBtn[0].click()
                        wait_float(5.5,7.5)
                        
                        driver.back()
                        break
            except:
                pass
        if ici >= forVal:
            break
    pg.alert('작업이 완료 되었습니다.')
    pg.alert('끝내시겠습니까?')
    exitApp()
    
 
    
    
    # user_mycafe_info
    
    # cafe-write-btn
    
    # menuLink226
    
    # cafe_main 프레임
    
    # #writeFormBtn
    
    # _noticeArticle
    
    # article-board tr
    
    # td_name
    
    # board-list
    
    
    # se-fs-
    
    
    # CommentWriter 






# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염

# 상품 들어가서 스크롤 내리고 나오기

def makeBlogContent():
    
    getInfoPostLink = 'https://m.blog.naver.com/overroad89/222696651375'

    page = requests.get(getInfoPostLink)
    soup = bs(page.text, "html.parser")
    elements = soup.select('.se-module.se-module-text')
    
    pg.alert(elements)
    
    
    allStr = []
    chkCount = 0
    for ele in elements:
        chkCount += 1
        p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
        chkResult = p.findall(str(ele))
        if chkCount == 1:
            subjectTemp = chkResult
        allStr = allStr + chkResult

    p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
    p_space = re.compile('\s\s')
    
    subjectArrTemp = []
    for sentence_s in subjectTemp:
        getStr = p_str.search(sentence_s)
        if getStr is None:
            subjectArrTemp.append(sentence_s)
    
    
    
    tempSubjectOn = ''
    for tempss in subjectArrTemp:
        if tempss == ' ':
            continue
        tempSubjectOn = tempSubjectOn + tempss
    
    
    
    subjectArr = tempSubjectOn.split(' ')
        
        

    for i in range(1, len(allStr)):
        for j, strin in enumerate(allStr):
            getStr = p_str.search(strin)
            if getStr is not None:
                allStr.pop(j)
                break
            getSpace = p_space.search(strin)
            if getSpace is not None:
                allStr.pop(j)
                break
            if strin == " ":
                allStr.pop(j)
                break
    allStr = "".join(allStr)
    
    # if len(allStr) < 1500:
    #     continue
    # if len(allStr) > 1900:
    #     sliceRanNum = random.randrange(1050, 1150)
    #     allStr = allStr[0:sliceRanNum]
    # break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    # driver.close()
    pg.alert(allContent)
    with open('./etc/text.txt', 'w') as f:
        f.write(allContent)
    
    exitApp()
















def makeContentArr(page):
    soup = bs(page.text, "html.parser")
    elements = soup.select('.se-module.se-module-text')
    sentenceEndArr = ['요','죠','다','용']
    
    allStr = []
    for ele in elements:
        p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
        chkResult = p.findall(str(ele))
        allStr = allStr + chkResult

    p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
    for i in range(1, len(allStr)):
        for j, strin in enumerate(allStr):
            getStr = p_str.search(strin)
            if getStr is not None:
                allStr.pop(j)
                break

            
    allStr = "".join(allStr)
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('   ', ' ')
    allStr = allStr.replace('  ', ' ')
    allStr = allStr.replace('  ', ' ')
    allStr = allStr.replace('  ', ' ')
    
    resetStrArr = allStr.split(' ')
    tempArr = []
    for i in range(0, len(resetStrArr)):
        for j, strin in enumerate(resetStrArr):
            if(strin != ''):
                if strin[-1] in sentenceEndArr:
                    chkArr = resetStrArr[0:j + 1]
                    tempArr.append(chkArr)
                    del resetStrArr[0:j + 1]
                    break
    
    return tempArr

def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp



def searchElement(ele):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element



def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        
        try:
            btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent




async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def mainToCafe():
    shs_item = searchElement('.shs_item')
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break
    
    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe')



    myCafeList = searchElement('.list_cafe__favorites li')
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")
    
    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break
        
    # 카페 진입 끝

def getBlogContent(subjectArr):
    # 블로그 글따기 시작

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get('https://section.blog.naver.com/BlogHome.naver')

    try:
        popup = driver.find_element(
            by=By.CSS_SELECTOR, value='#floatingda_home')
        popupClostBtn = popup.find_elements(by=By.CSS_SELECTOR, value='button')
        popupClostBtn[-1].click()
    except:
        pass

    while True:
        
        
        try:
            nCategoryList = driver.find_elements(by=By.CSS_SELECTOR, value='.navigator_category a')
            categoryRanVal = random.randrange(0, len(nCategoryList) - 1)
            nCategoryList[categoryRanVal].click()

            wait_float(0.5, 0.9)

            paginationNum = driver.find_elements(by=By.CSS_SELECTOR, value='.pagination span')
            driver.execute_script("arguments[0].scrollIntoView();", paginationNum[0])
            paginationRanVal = random.randrange(0, len(paginationNum) -1)
            getClickPage = paginationNum[paginationRanVal].find_element(by=By.CSS_SELECTOR, value='a')
            getClickPage.click()
            wait_float(0.5, 0.9)

            infoPostList = driver.find_elements(by=By.CSS_SELECTOR, value='.info_post')
            infoPostRanVal = random.randrange(0, len(infoPostList) - 1)
            getInfoPostTag_a = infoPostList[infoPostRanVal].find_element(by=By.CSS_SELECTOR, value='.desc a')
            getInfoPostLink = getInfoPostTag_a.get_attribute('href')
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')
        except:
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
            wait_float(2.5,3.5)
            continue

        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')

        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        allStr = "".join(allStr)
        if len(allStr) < 600:
            continue
        if len(allStr) > 1200:
            sliceRanNum = random.randrange(1050, 1150)
            allStr = allStr[0:sliceRanNum]
        break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    driver.close()
    return allContent


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
