
import random
import threading
import time
from datetime import date, datetime, timedelta
from dateutil.relativedelta import *
import sys
import os
from pathlib import Path
from typing import Optional
from pyparsing import And
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import aiohttp
import asyncio


"""
주요 변수
nowAction ( write / reply ) / 글쓰는지 댓글인지 체크
nowWriteStatus (optimize / basic ) / 최적화인지 아닌지 체크
allCount : 전체 변수, 글 / 댓글 나누기 위함
writeCount : 글쓰기 변수 nowAction 이 write일때 하나씩 증가, 최적화 아이디인지 판별하기 위함

cafe_id.cell(세로(열), 가로(행)).value
"""

###


def goScript(getDict):
    
    # chromeVersionChkPath = 'C:\\Users\\pcy\\AppData\\Local\\Google\\Chrome\\User Data\\default'
    chromeVersionChkPath = 'C:\\Users\\드림모어\\AppData\\Local\\Google\\Chrome\\User Data\\Default'
    
    global driver
    
    

    cafe_optimize_file = load_workbook('./etc/naver_optimiz.xlsx')
    cafe_optimize = cafe_optimize_file.active
    cafe_id_file = load_workbook('./etc/naver_id.xlsx')
    cafe_id = cafe_id_file.active

    allCount = 0
    writeCount = 0
    endOptimize = ''
    nowAction = ''
    preIp = ''
    chk_extesion = ['jpg', 'jpeg', 'JPG', 'png', 'PNG', 'gif']

    with open('./etc/cafe_info.txt', 'r') as f:
        cafeAllInfo = f.readlines()

    for i in range(0, len(cafeAllInfo)):
        cafeAllInfo[i] = cafeAllInfo[i].replace('\n', '')

    print(cafeAllInfo)

    pg.alert(text='시작 대기!!')

    cafeName = cafeAllInfo[1]
    boardListKor = cafeAllInfo[2].split(',')
    boardListNum = cafeAllInfo[3].split(',')

    while True:
        
        
        # 시작
        allCount += 1
        
        # 아이피 변경
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)
        while True:
            getIP = changeIpSpeed()
            print(getIP)
            if not preIp == getIP:
                preIp = getIP
                break

        # 5로 나누어서 나머지가 1이면 (5의 배수 + 1 값이면 글쓰기 진행)
        print('아이피 변경 완료')
        nowActionNum = allCount % 4
        # nowActionNum = allCount % 2
        if nowActionNum == 1:
            nowAction = 'write'
            writeCount += 1
        else:
            nowAction = 'reply'
            nowWriteStatus = ""

        print(f'일단 현재 작업은? {nowAction}')
        # 최적화 아이디인지 일반 아이디인지 체크
        
        
        # 첫번째 란이 비어있거나 / 4번째란이 채워져 있으면 최적화 작업 끝
        if endOptimize == '' and nowAction == 'write':
            if os.path.exists(f'./etc/content/id_{writeCount}'):
                optimizeChkVal1 = cafe_optimize.cell(writeCount, 2).value
                optimizeChkVal2 = cafe_optimize.cell(writeCount, 4).value
                if optimizeChkVal1 is None or optimizeChkVal2 is not None:
                    endOptimize = 'on'
                    nowWriteStatus = 'basic'
                else:
                    nowWriteStatus = 'optimize'
            else:
                endOptimize = 'on'
                nowWriteStatus = 'basic'
        else:
            nowWriteStatus = 'basic'

        print(f'최적화 여부는?? {nowWriteStatus}')

        # 최적화 글쓰기 / 일반 글쓰기 / 댓글쓰기 각 정보 (크롬정보 / 아이디 / 비번 / 게시판 번호 등) 부여하기
        if nowWriteStatus == 'optimize' and nowAction == 'write':
            # 최적화 아이디 일때
            uaSet = cafe_optimize.cell(writeCount, 1).value
            if uaSet is None:
                uaSet = getUaNum()
                cafe_optimize.cell(writeCount, 1).value = uaSet
                cafe_optimize_file.save('./etc/naver_optimiz.xlsx')

            nId = cafe_optimize.cell(writeCount, 2).value
            nPwd = cafe_optimize.cell(writeCount, 3).value
            nBoardName = cafe_optimize.cell(writeCount, 6).value
            nBoardNum = cafe_optimize.cell(writeCount, 7).value

            cafe_optimize.cell(writeCount, 4).value = datetime.now()
            cafe_optimize_file.save('./etc/naver_optimiz.xlsx')

        elif nowWriteStatus == 'basic' or nowAction == 'reply':
            # 일반 글쓰기 or 댓글 쓰기 일때 (안써진거 or 쓴지 3일 지난거)
            # 먼저 엑셀에서 사용한지 3일이 지난 값 가지고 오기
            nidExLength = getExLength(cafe_id)
            chkArr = asyncio.run(getEmptyArr(nidExLength, cafe_id))
            getRanVal = random.randrange(0, len(chkArr))
            getRanWorkVal = chkArr[getRanVal]

            uaSet = cafe_id.cell(getRanWorkVal, 1).value
            if uaSet is None:
                uaSet = getUaNum()
                cafe_id.cell(getRanWorkVal, 1).value = uaSet
                cafe_id_file.save('./etc/naver_id.xlsx')

            nId = cafe_id.cell(getRanWorkVal, 2).value
            nPwd = cafe_id.cell(getRanWorkVal, 3).value
            boardGetRan = random.randrange(0, 2)
            nBoardName = boardListKor[boardGetRan]
            nBoardNum = boardListNum[boardGetRan]

            print(datetime.now().date())
            cafe_id.cell(getRanWorkVal, 4).value = datetime.now()
            cafe_id_file.save('./etc/naver_id.xlsx')

        # 테스트겸 냅두자
        try:
            getVal = getRanWorkVal
        except:
            getVal = writeCount

        print(f'{getVal}번째 있는 아이디로 {nowWriteStatus} {nowAction}작업, 크롬 정보 : {uaSet} / 아이디 : {nId} / 비번 : {nPwd} / 게시판 이름 {nBoardName}')

        print('정보 얻기 완료')
        
        
        # 일반 아이디 글쓰기 작업
        if nowAction == 'write' and nowWriteStatus == 'basic':

            blog_start = time.time()

            # 블로그 글따기 시작!!!
            # 엑셀로 랜덤 돌려서 제목 뽑기
            cafe_ex_file = load_workbook('./etc/subject_list.xlsx')
            cafe_ex = cafe_ex_file.active

            subjectCountArr = []
            for i in range(1, 5):
                k = 0
                while True:
                    k += 1
                    chkVal = cafe_ex.cell(k, i).value
                    if chkVal is None:
                        subjectCountArr.append(k)
                        break

            subjecArr = []
            for i, subjectCount in enumerate(subjectCountArr):
                if i == 1 or i == 2:
                    passNum = random.randrange(1, 3)
                    if passNum != 1:
                        continue
                getConNum = random.randrange(1, subjectCount)
                chkVal = cafe_ex.cell(getConNum, i+1).value
                subjecArr.append(chkVal)

            print('제목 생성 완료')
            # 엑셀로 랜덤 돌려서 제목 뽑기 끝 이제 아래 블로그 컨텐츠 생성 함수에 넣고 막글 뽑자!
            
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service)
            print('블로그 글 따기 시작~~~~~~~~~~~~~~~~~~~~~~~~~!!!!!!')
            blog_content = getBlogContentChrome(subjecArr)
            
            if blog_content == 'errRobot':
                print('에러가 났어요!!! 처음으로 돌아가야해요!!')
                allCount = allCount - 1
                continue
            
            print('블로그 글 따기 완료')
            subject = " ".join(subjecArr)
            with open("./etc/content/write_content.txt", "w") as f:
                f.write(subject)
                f.write('\n')
                f.write(blog_content)
            # 블로그 글따기 끝!!
            
            # 모바일 버전 일반 글쓰기 시작~~~~
            nLoginTimeStart = time.time()

            # 네이버 메인에서 카페 진입 시작!
            with open(f'./etc/useragent/useragent_all.txt') as f:
                ua_data = f.readlines()[uaSet]
                ua_data = ua_data.replace('\n', '')

            options = Options()
            user_agent = ua_data
            options.add_argument('user-agent=' + user_agent)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(chrome_options=options, service=service)

            driver.get('https://www.naver.com')
            
            
            errchk = naverLogin(nId, nPwd)
            if errchk is not None:
                cafe_id.cell(getRanWorkVal, 4).value = errchk
                cafe_id_file.save('./etc/naver_id.xlsx')
                allCount = allCount - 1
                driver.close()
                continue
            
            # 모바일 카페 글쓰기
            mobileCafeWrite(cafeName,nBoardName,chk_extesion)
            
            goToHome = searchElement('.header h1')
            untilEleGone(goToHome[0], '.post_title')
            mobileReplyAction()
            
        
        
        # 최적화 아이디 글쓰기 작업
        if nowAction == 'write' and nowWriteStatus == 'optimize':
            options = Options()
            user_data = chromeVersionChkPath
            service = Service(ChromeDriverManager().install())
            options.add_argument(f"user-data-dir={user_data}")
            options.add_argument(f'--profile-directory={uaSet}')
            driver = webdriver.Chrome(service=service, chrome_options=options)
            driver.set_window_size(1180, 910)
            driver.set_window_position(0,0) 
            
            
            driver.get('https://www.naver.com')
            
            # 네이버 로그인~~~~~~
            
            
            loginBtn = searchElement('.sc_login')
            loginBtn[0].click()
            
            searchElement('#id')
            focus_window('chrome')
            wait_float(0.3,0.9)
            while True:
            
                pg.click(50,400)
                inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
                inputId.click()
                wait_float(0.3,0.9)
                cb.copy(nId)
                wait_float(0.3,0.9)
                pg.hotkey('ctrl', 'a')
                wait_float(0.3,0.9)
                pg.hotkey('ctrl', 'v')
                inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
                if inputId.get_attribute('value') != "":
                    break
            
            while True:
                inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
                inputPw.click()
                wait_float(0.3,0.9)
                cb.copy(nPwd)
                wait_float(0.3,0.9)
                pg.hotkey('ctrl', 'a')
                wait_float(0.3,0.9)
                pg.hotkey('ctrl', 'v')
                inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
                if inputPw.get_attribute('value') != "":
                    break
            
            btnLogin = searchElement('.btn_login')
            btnLogin[0].click()
            
            
            while True:
                try:
                    WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
                    break
                except:
                    driver.get('https://www.naver.com')
            
            navItem = searchElement('.nav_item')
            for mitem in navItem:
                if mitem.text == '카페':
                    mitem.click()
                    break
            
            cafeList = searchElement('.user_mycafe_info')
            getInCafe = ""
            for cafeOn in cafeList:
                if cafeAllInfo[0] in cafeOn.text:
                    cafeOn.click()
                    getInCafe = "on"
                    break
            
            if getInCafe == "":
                driver.get(cafeAllInfo[1])
            else:
                driver.switch_to.window(driver.window_handles[0])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            
            searchElement('.cafe-write-btn')
            
            workBoardLink = searchElement(f'#menuLink{nBoardNum}')
            workBoardLink[0].click()
            
            
            driver.switch_to.frame('cafe_main')
            
            # 카페에 글 작성하기
            
            workBoardWriteBtn = searchElement('#writeFormBtn')
            workBoardWriteBtn[0].click()
            
            wait_float(1.5,2.5)
            driver.switch_to.window(driver.window_handles[1])
            
            with open(f'./etc/content/id_{writeCount}/content.txt', 'r') as f:
                getContents = f.readlines()
            
            subjectArea = searchElement('.FlexableTextArea')
            subjectArea[0].click()
            keyboard.write(text=getContents[0], delay=0.05)
            wait_float(0.8,1.9)
            
            textArea = searchElement('.se-content')
            textArea[0].click()
            for i, getline in enumerate(getContents):
                if i == 0:
                    continue
                getline = getline.replace('\n', '')
                getImgAction = getline.split('|')
                if getline == 'enter':
                    pg.press('enter')
                    wait_float(0.5, 0.9)
                elif getImgAction[0] == 'img_line':
                    imgBtn = searchElement('.se-document-toolbar li')
                    imgBtn[0].click()
                    nowPath = os.getcwd()
                    imagePath = nowPath + f"\etc\content\id_{writeCount}"
                    wait_float(1.5, 2.2)
                    pyperclip.copy(imagePath)
                    wait_float(0.5, 0.9)
                    pg.hotkey('ctrl', 'v')
                    wait_float(0.5, 0.9)
                    pg.press('enter')
                    
                    wait_float(0.9, 1.6)
                    pyperclip.copy(getImgAction[1])
                    wait_float(0.5, 0.9)
                    pg.hotkey('ctrl', 'v')
                    wait_float(0.5, 0.9)
                    pg.press('enter')
                    
                    wait_float(3.5, 4.8)
                else:
                    keyboard.write(text=getline, delay=0.05)
                    wait_float(0.5, 0.9)
                    pg.press('enter')
            BaseButton = searchElement('.BaseButton')
            BaseButton[0].click()
            
            wait_float(1.5,2.5)
            driver.switch_to.frame('cafe_main')
            
            BaseButton = searchElement('.button_url')
            BaseButton[0].click()
            wait_float(0.5,0.9)
            getLinkData = cb.paste()
            wait_float(0.5,0.9)
            
            with open(f'./etc/content/id_{writeCount}/reply.txt', 'r') as f:
                getTempReplys = f.readlines()
                getTempReplys.insert(0, '0\n')
                getTempReplysName_temp = getLinkData.split('/')
                getTempReplysName = getTempReplysName_temp[-1]
                
            with open(f'./etc/content/temp_reply/{getTempReplysName}.txt', 'w') as f:
                f.writelines(''.join(getTempReplys))
            
            
            with open('./etc/work_link.txt', 'a') as f:
                f.write('\n')
                f.write(getLinkData)

            with open('./etc/work_link.txt', 'r') as f:
                chkLines = f.readlines()

            if len(chkLines) > 6:
                setLines = chkLines[-6:]
                
            linkContent = ''
            for linkline in setLines:
                linkContent = linkContent + linkline
                
            with open('./etc/work_link.txt', 'w') as w:
                w.write(linkContent)
                
            # worklink 설정 끝~~~~~~~~~~~
                
            wait_float(0.5,0.9)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.frame('cafe_main')
            
            # 글쓰기 및 worklink에 링크 추가 완료 PC버전 댓글 달기 GO!
            with open(f'./etc/work_link.txt') as f:
                workLinkList = f.readlines()
                
            for i, workLink in enumerate(workLinkList):
                driver.switch_to.default_content()
                wait_float(0.3,0.9)
                workBoardLink = searchElement('#menuLink0')
                workBoardLink[0].click()
                
                workLink_temp = workLink.replace('\n', '')
                workLinkOn = workLink_temp.split('/')[-1]
                
                driver.switch_to.frame('cafe_main')
                articleDiff = searchElement('.article-board')
                articleList = articleDiff[1].find_elements(by=By.CSS_SELECTOR, value=".td_article")
                for uu, article in enumerate(articleList):
                    getArticleHref = article.find_element(by=By.CSS_SELECTOR, value=".article").get_attribute('href')
                    getArticleHref = getArticleHref.split('id=')[-1].split('&')[0]
                    if workLinkOn == getArticleHref:
                        
                        while True:
                            tempArticleWrap = searchElement('.article-board')
                            articleVal = tempArticleWrap[1].find_elements(by=By.CSS_SELECTOR, value=".td_article .article")
                            articleVal[uu].click()
                            wait_float(1.5,2.2)
                            try:
                                driver.find_element(by=By.CSS_SELECTOR, value=".article_header")
                                break
                            except:
                                continue
                        pg.moveTo(200, 500)
                        randomFor = random.randrange(3, 6)
                        for i in range(1, randomFor):
                            wait_float(1.5, 2.5)
                            pg.scroll(-500)
                        
                        randomActionVal = random.randrange(1, 4)
                        if randomActionVal == 1:
                            getReply = ""
                            if os.path.exists(f'./etc/content/temp_reply/{workLinkOn}.txt'):
                                with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'r') as f:
                                    getTempReplys = f.readlines()
                                    getTempNum = getTempReplys[0].replace('\n', '')
                                    try:
                                        getReply = getTempReplys[int(getTempNum) + 1]
                                        getTempReplys[0] = str(int(getTempNum) + 1) + '\n'
                                    except:
                                        getReply = ""
                                wait_float(0.2, 0.7)
                                if getReply != "":
                                    with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'w') as f:
                                        f.writelines(''.join(getTempReplys))
                            wait_float(0.2, 0.7)
                            if getReply == "":
                                with open(f'./etc/all_reply.txt', 'r') as f:
                                    getTempReplysAll = f.readlines()
                                    getTempRanNum = random.randrange(
                                        0, len(getTempReplysAll))
                                    getReply = getTempReplysAll[getTempRanNum]
                                    
                            wait_float(1.5,2.8)
                            try:
                                replyArea = driver.find_element(by=By.CSS_SELECTOR, value=".comment_inbox_text")
                                replyArea.click()
                                keyboard.write(text=getReply, delay=0.03)
                                wait_float(1.5,2.5)
                                replySuccessBtn = searchElement('.btn_register')
                                driver.execute_script("arguments[0].scrollIntoView();", replySuccessBtn[0])
                                replySuccessBtn[0].click()
                                wait_float(2.5,3.5)
                                print(replyArea)
                            except:
                                pass
                        break

        # ★★★★★★★★ 댓글 작성 시작!!
        # 댓글달기 로그인
        if nowAction == 'reply':
            with open(f'./etc/useragent/useragent_all.txt') as f:
                ua_data = f.readlines()[uaSet]
            options = Options()
            user_agent = ua_data
            options.add_argument('user-agent=' + user_agent)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(chrome_options=options, service=service)

            nLoginTimeStart = time.time()

            driver.get('https://www.naver.com')

            errchk = naverLogin(nId, nPwd)
            if errchk is not None:
                if nowWriteStatus == 'basic':
                    cafe_id.cell(getRanWorkVal, 4).value = errchk
                    cafe_id_file.save('./etc/naver_id.xlsx')
                elif nowWriteStatus == 'optimize':
                    cafe_optimize.cell(writeCount, 4).value = errchk
                    cafe_optimize_file.save('./etc/naver_optimiz.xlsx')
                allCount = allCount - 1
                driver.close()
                continue

            driver.get(cafeName)
            mobileReplyAction()
            
        # 카페 메인 진입 끝! 게시글 클릭 시작!
        
        
        nowWriteStatus = ''
        driver.close()
    


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염

# 상품 들어가서 스크롤 내리고 나오기




def naverLogin(load_id, load_pass):
    search_bar = searchElement(".sch_ico_aside")
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap")
    login_btn[0].click()

    # 로그인 부분

    focus_window('로그인')
    while True:
        searchElement("#id")

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header")

        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass
        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".action_inner")
            if unPwd:
                return "비정상적 활동"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close")
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp


def mobileCafeWrite(cafeName,nBoardName,chk_extesion):
    driver.get(cafeName)
    print('카페 진입 완료')

    writeBtn = searchElement('.inner_box .btn_write')
    untilEleGone(writeBtn[0], '.inner_box .btn_write')

    # 메뉴 선택하기
    selectBox = searchElement('.selectbox')
    untilEleShow(selectBox[0], '.layer_dimmed')
    selBoard = searchElement('.select_board li')
    for board in selBoard:
        if board.text == nBoardName:
            board.click()
            break

    print('메뉴 선택 완료')

    # 메뉴 선택 완료! 글쓰기 시작~~~~
    with open("./etc/content/write_content.txt", "r") as f:
            getContents = f.readlines()

    chkImg = [0, 0, 0, 0, 0, 0]

    # 글 작성 전 창 체크
    focus_window('글쓰기')

    articleWriteFormSubject = searchElement('.ArticleWriteFormSubject')
    articleWriteFormSubject[0].click()
    keyboard.write(text=getContents[0], delay=0.3)

    oneEditor = searchElement('#one-editor')
    oneEditor[0].click()
    for i, line in enumerate(getContents):
        if i == 0:
            continue
        
        try:
            line_temp = line.replace('\n', '')
            chkImg = line_temp.split('|')
        except:
            pass
        
        line = line.replace('\n', '')
        
        if chkImg[0] == 'img_line':
            imageUpload = searchElement('.se-toolbar-item-image')
            imageUpload[0].click()

            nowPath = os.getcwd()
            imagePath = nowPath + "\etc\content\images"
            imageList = os.listdir(imagePath)

            while True:
                getImage = imageList[random.randrange(0, len(imageList))]
                getImage_ex = getImage.split('.')[-1]
                if getImage_ex in chk_extesion:
                    break
            wait_float(1.5, 2.2)
            pyperclip.copy(imagePath)
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl', 'v')
            wait_float(0.5, 0.9)
            pg.press('enter')

            wait_float(0.9, 1.6)
            pyperclip.copy(getImage)
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl', 'v')
            wait_float(0.5, 0.9)
            pg.press('enter')

            wait_float(3.5, 4.8)
            # 끝난다음 초기화
            chkImg = [0, 0, 0, 0, 0, 0]
        elif line == 'enter':
            pg.press('enter')
        else:
            keyboard.write(text=line, delay=0.05)
            keyboard.write(text='\n', delay=0.05)
        wait_float(0.7, 1.3)
        
    successBtn = searchElement('.GnbBntRight__green')
    untilEleGone(successBtn[0], '.GnbBntRight__green')

    print('글쓰기 완료')

    getLinkMore = searchElement('.btn_aside .more')
    getLinkMore[0].click()

    getLink = searchElement('.layer_list li')
    getLink[3].click()

    getLinkData = cb.paste()
    
    with open('./etc/work_link.txt', 'a') as f:
        f.write('\n')
        f.write(getLinkData)

    with open('./etc/work_link.txt', 'r') as f:
        chkLines = f.readlines()

    if len(chkLines) > 6:
        setLines = chkLines[-6:]
        
    linkContent = ''
    for linkline in setLines:
        linkContent = linkContent + linkline
        
    with open('./etc/work_link.txt', 'w') as w:
        w.write(linkContent)




def mobileReplyAction():
    with open(f'./etc/work_link.txt') as f:
            workLinkList = f.readlines()
    
    print('문제점22222')
    for i, workLink in enumerate(workLinkList):
        
        workLink_temp = workLink.replace('\n', '')
        workLinkOn = workLink_temp.split('/')[-1]
        boardListAll = searchElement('.list_area li')

        for boardList in boardListAll:
            chkBoardLink = boardList.find_element(by=By.CSS_SELECTOR, value='.txt_area').get_attribute('href')
            if workLinkOn in chkBoardLink:
                untilEleGone(boardList, '.txt_area')
                pg.moveTo(300, 500)
                randomFor = random.randrange(2, 5)
                for i in range(1, randomFor):
                    wait_float(1.5, 2.5)
                    pg.scroll(-500)
                break

        # 게시글 클릭 완료 댓글 쓰기 시작!

        randomActionVal = random.randrange(1, 4)
        print(randomActionVal)
        if randomActionVal == 1:
            
            # 댓글 작성 전 창 체크
            replyGoBtn = searchElement('.f_reply')
            untilEleShow(replyGoBtn[0], '.HeaderIcon')

            wait_float(0.5, 0.9)
            replyBtn = searchElement('.comment_textarea')
            replyBtn[0].click()

            getReply = ""
            if os.path.exists(f'./etc/content/temp_reply/{workLinkOn}.txt'):

                with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'r') as f:
                    getTempReplys = f.readlines()
                    getTempNum = getTempReplys[0].replace('\n', '')
                    try:
                        getReply = getTempReplys[int(getTempNum) + 1]
                        getTempReplys[0] = str(int(getTempNum) + 1) + '\n'
                    except:
                        getReply = ""
                wait_float(0.2, 0.7)
                if getReply != "":
                    with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'w') as f:
                        f.writelines(''.join(getTempReplys))
            wait_float(0.2, 0.7)
            if getReply == "":
                with open(f'./etc/all_reply.txt', 'r') as f:
                    getTempReplysAll = f.readlines()
                    getTempRanNum = random.randrange(
                        0, len(getTempReplysAll))
                    getReply = getTempReplysAll[getTempRanNum]

            focus_window('카페')
            keyboard.write(text=getReply, delay=0.1)
            replySuccessBtn = searchElement('.btn_done')
            untilEleGone(replySuccessBtn[0], '.btn_done')

            goToPostiongBtn = searchElement('.HeaderGnbLeft')
            untilEleGone(goToPostiongBtn[0], '.HeaderGnbLeft')

        goToHome = searchElement('.header h1')
        untilEleGone(goToHome[0], '.post_title')


def changeIpSpeed():
    
    while True:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        
        try:
            ondevice = device[0]
            ondevice.shell("input keyevent KEYCODE_POWER")
            ondevice.shell("svc data disable")
            ondevice.shell("settings put global airplane_mode_on 1")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

            ondevice.shell("svc data enable")
            ondevice.shell("settings put global airplane_mode_on 0")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
            time.sleep(3)
        except:
            pass
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue

        driver.get('https://fast.com/ko/')
        searchElement('.speed-results-container')
        time.sleep(3)
        getInternetRapidEle = searchElement('.speed-results-container')
        getInternetRapid = getInternetRapidEle[0].text
        if float(getInternetRapid) < 2.7:
            continue
        else:
            driver.close()
            break

    return getIp


def searchElement(ele):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element



def untilEleShow(clickEle, searchEle):
    while True:
        print('정체~~~~~~~')
        print(clickEle.text)
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


# def untilEleGone(clickEle, searchEle):
#     while True:
#         try:
#             clickEle.click()
#             time.sleep(1)
#             try:
#                 btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
#                 if btnEle is None:
#                     return
#             except:
#                 return
#         except:
#             pass

def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass

        try:
            btnEle = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent


async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet


def mainToCafe():
    shs_item = searchElement('.shs_item')
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break

    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe')

    myCafeList = searchElement('.list_cafe__favorites li')
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")

    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(
            by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break

    # 카페 진입 끝


def getBlogContentChrome(subjectArr):
   
    with open('./etc/find_keyword.txt', 'r') as r:
        allKeyword = r.readlines()
        
    driver.get('https://www.google.com/')
    
    
    searchCount = 0
    while True:
        searchCount += 1
        keyCount = random.randrange(0, len(allKeyword))
        getKeyword = allKeyword[keyCount]
        getKeyword = getKeyword.replace('\n', '')
        searchBar = searchElement('.gLFyf.gsfi')
        wait_float(0.5,0.8)
        searchBar[-1].click()
        wait_float(0.5,0.8)
        pg.hotkey('ctrl', 'a')
        wait_float(0.5,0.8)
        keyboard.write(text=f'site:blog.naver.com {getKeyword}', delay=0.5)
        pg.press('enter')
        wait_float(0.8,1.5)
        pg.press('enter')
        
        print('검색 완료~~~~~~~~~~~~~~~~~~~!!!!!!')
        nowpage = 0
        if searchCount == 1:
            getTools = searchElement('.t2vtad')
            getTools[0].click()
            wait_float(0.3,0.9)
            
            getToolsIf = searchElement('.KTBKoe')
            for getToolsIfOn in getToolsIf:
                if '날짜' in getToolsIfOn.text:
                    getToolsIfOn.click()
            wait_float(0.3,0.9)
                    
            getPeriodIf = searchElement('.y0fQ9c')
            for getPeriodIfOn in getPeriodIf:
                if '설정' in getPeriodIfOn.text:
                    getPeriodIfOn.click()
            wait_float(0.3,0.9)
            
            today = datetime.today()
            print(today)
            
            before_one_year = today - relativedelta(years=3)

            # this_month_first = datetime(before_one_year.year, before_one_year.month, 1)
            # chkMonthFirst = this_month_first.strftime('%m/%d/%Y')
            # driver.find_element(by=By.CSS_SELECTOR, value='.OouJcb').send_keys(chkMonthFirst)
            # next_month = datetime(before_one_year.year, before_one_year.month, 1) + relativedelta(months=1)
            # this_month_last = next_month + relativedelta(seconds=-1)
            # chkMonthLast = this_month_last.strftime('%m/%d/%Y')
            
            chkThreeyearAgo = before_one_year.strftime('%m/%d/%Y')
            driver.find_element(by=By.CSS_SELECTOR, value='.OouJcb').send_keys(chkThreeyearAgo)
            driver.find_element(by=By.CSS_SELECTOR, value='.rzG2be').send_keys(chkThreeyearAgo)
            wait_float(0.8,1.5)
            pg.press('enter')
        print('기간 설정 완료~~~~~~~~~~~~~~~~~~~!!!!!!')
            
        try:
            driver.find_element(by=By.CSS_SELECTOR, value='.NVbCr')
        except:
            continue
        getPagingList = searchElement('.NVbCr')
        if len(getPagingList) < 2:
            continue
        getPgCount = random.randrange(0,len(getPagingList))
        if getPgCount != nowpage:
            nowpage = getPgCount
            getPagingList[getPgCount].click()
            
        try:
            getBlogLink = driver.find_elements(by=By.CSS_SELECTOR, value='.yuRUbf')
        except:
            print('문제 생긴거 맞나요?!?!?!')
            return 'errRobot'

        getBlogLinkCount = random.randrange(0,len(getBlogLink))
        
        getInfoPostLink = getBlogLink[getBlogLinkCount].find_element(by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if '//m.' not in getInfoPostLink:
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')
            
            
        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')
        print(getInfoPostLink)
        print('블로그 파싱 완료! 글 다듬기 시작~~~~~~~~~~~~~~~~~~~!!!!!!')
        # se-title-text 제목
        
        try:
            getSubjectEle = soup.select('.se-title-text')
            getSubject = str(getSubjectEle[0])
            
            getSubject = getSubject.split('-->')[-2].replace('','').split('<!--')[0]
            getSubject = re.sub(r"[^\uAC00-\uD7A3\s]", "", getSubject)
            getSubjectArr = getSubject.replace('  ', ' ').split(' ')

            allStr = []
            for ele in elements:
                p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
                chkResult = p.findall(str(ele))
                allStr = allStr + chkResult

            p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
            p_space = re.compile('\s\s')

            for i in range(1, len(allStr)):
                for j, strin in enumerate(allStr):
                    getStr = p_str.search(strin)
                    if getStr is not None:
                        allStr.pop(j)
                        break
                    getSpace = p_space.search(strin)
                    if getSpace is not None:
                        allStr.pop(j)
                        break
                    if strin == " ":
                        allStr.pop(j)
                        break
            
            getAllStr = []
            for strOn in allStr:
                addVal = ''
                for tempsub in getSubjectArr:
                    if len(tempsub) < 2:
                        continue
                    if tempsub in strOn:
                        addVal = ''
                        break
                    addVal = 'on'
                
                if addVal == 'on':
                    getAllStr.append(strOn)
            
            
            
            
            allStr = " ".join(getAllStr)
            
            if len(allStr) < 400:
                continue
            if len(allStr) > 600:
                sliceRanNum = random.randrange(400, 600)
                allStr = allStr[0:sliceRanNum]
                
            # 제목에 들어간 단어들 삭제하기
            break
        except:
            continue
    print('잡다한거 (태그, 영어, 숫자 등) 빼기 완료 ~~~~~~~~~~~~~~~~~~~!!!!!!')

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)
        
        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 4)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass
    print('마무리 첫번째 체크~~~~~~~~~~~~~~~~~~~!!!!!!')
    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
            insubject = " ".join(subjectArr)
            allContent = allContent + insubject + '\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        
        if len(resetListArr)-1 != i:
            allContent = allContent + '\n'
        
    allContent = allContent + subjectArr[-1]
    print('마무리 완료~~~~~~~~~~~~~~~~~~~!!!!!!')
    driver.close()
    return allContent
        
    
    
    
    
    # #시작일
    # OouJcb
    
    # #종료일
    # rzG2be
    
    
    
    
    
    
    
    
    
    
    
     
    

def getBlogContent(subjectArr):
    # 블로그 글따기 시작

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get('https://section.blog.naver.com/BlogHome.naver')

    try:
        popup = driver.find_element(
            by=By.CSS_SELECTOR, value='#floatingda_home')
        popupClostBtn = popup.find_elements(by=By.CSS_SELECTOR, value='button')
        popupClostBtn[-1].click()
    except:
        pass

    while True:

        try:
            nCategoryList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.navigator_category a')
            categoryRanVal = random.randrange(0, len(nCategoryList) - 1)
            nCategoryList[categoryRanVal].click()

            wait_float(0.5, 0.9)

            paginationNum = driver.find_elements(
                by=By.CSS_SELECTOR, value='.pagination span')
            driver.execute_script(
                "arguments[0].scrollIntoView();", paginationNum[0])
            paginationRanVal = random.randrange(0, len(paginationNum) - 1)
            getClickPage = paginationNum[paginationRanVal].find_element(
                by=By.CSS_SELECTOR, value='a')
            getClickPage.click()
            wait_float(0.5, 0.9)

            infoPostList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.info_post')
            infoPostRanVal = random.randrange(0, len(infoPostList) - 1)
            getInfoPostTag_a = infoPostList[infoPostRanVal].find_element(
                by=By.CSS_SELECTOR, value='.desc a')
            getInfoPostLink = getInfoPostTag_a.get_attribute('href')
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')
        except:
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
            wait_float(2.5, 3.5)
            continue

        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')

        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        allStr = "".join(allStr)
        if len(allStr) < 600:
            continue
        if len(allStr) > 1200:
            sliceRanNum = random.randrange(1050, 1150)
            allStr = allStr[0:sliceRanNum]
        break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    driver.close()
    return allContent


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
