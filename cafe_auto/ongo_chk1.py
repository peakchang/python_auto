import linecache
import random
import threading
import time
import sys
import os

import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import random


def goScript():



    for p in range(1, 100):
        changeIp()
        wb = load_workbook('네이버아이디.xlsx')
        id_excel = wb.active

        set_ran = random.randrange(1, 500)
        load_id = id_excel.cell(set_ran, 2).value
        load_pass = id_excel.cell(set_ran, 3).value
        wb.save('네이버아이디.xlsx')



        ua_data = linecache.getline('./etc/useragent/useragent_all.txt', random.randrange(1, 14)).strip()
        get_line = random.randrange(0, 3)
        with open("post_link.txt") as postlink:
            link_var = postlink.readlines()[get_line]

        options = Options()
        user_agent = ua_data
        options.add_argument('user-agent=' + user_agent)
        global driver
        driver = webdriver.Chrome(chrome_options=options)

        driver.get('https://www.naver.com')


        start = time.time()
        search_bar = searchElement(".sch_ico_aside")
        end = time.time()

        print(f"{end - start:.5f} sec")
        pg.alert(text='대기요!!!!')



        # 로그인 부분
        search_bar = searchElement(".sch_ico_aside")
        print(search_bar)
        search_bar[0].click()
        login_btn = searchElement(".ss_profile_wrap")
        login_btn[0].click()

        searchElement("#id")
        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'a')
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'v')
        wait_float(0.5, 1.0)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'a')
        wait_float(0.5, 1.0)
        pg.hotkey('ctrl', 'v')
        wait_float(0.5, 1.0)
        id_input_value = id_input[0].get_attribute('value')
        pg.hotkey('enter')

        time.sleep(5)
        # 로그인 부분 끝
        driver.get('https://www.naver.com')

        searchElement('#MM_logo')
        time.sleep(2)
        driver.execute_script(f'''
                    const logo_area = document.querySelector('#MM_logo');
                    let __addtag = document.createElement("div");
                    __addtag.innerHTML = "<div><a href='{link_var}' class='setbliiiii'>양양 파르나스</a></div>";
                    logo_area.appendChild(__addtag);
                    let tagArea = document.getElementById('tagArea');
            ''');

        time.sleep(2)
        make_link = driver.find_element(by=By.CSS_SELECTOR, value=".setbliiiii")
        print(make_link)
        time.sleep(2)
        make_link.click()
        time.sleep(8)
        chk_links = driver.find_elements(by=By.CSS_SELECTOR, value="a")

        print(chk_links)
        for chk_link in chk_links:
            print(chk_link.text)
            if chk_link.text == '양양 파르나스':
                driver.execute_script("arguments[0].scrollIntoView();", chk_link)
                time.sleep(2)
                pg.hotkey('up')
                time.sleep(1)
                pg.hotkey('up')
                time.sleep(1)
                pg.hotkey('up')
                time.sleep(1)

                chk_link.click()

        time.sleep(150)





def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell("am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell("am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(5)
        getIp = requests.get("http://ip.jsontest.com").json()['ip']
    except:
        time.sleep(5)
        getIp = requests.get("http://ip.jsontest.com").json()['ip']
    return getIp


def searchElement(ele):
    time.sleep(1)
    re_count = 1
    element = ""
    while True:
        if re_count % 5 == 0:
            print("새로고침!!!!")
            driver.refresh()
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            re_count += 1

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.5, 1.2)
    return selected_element


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winname: str):
    win_list = gw.getAllTitles()
    get_list = [list for list in win_list if winname in list]
    win = gw.getWindowsWithTitle(get_list[0])[0]
    win.activate()
