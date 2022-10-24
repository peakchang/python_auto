
import random
import threading
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
from typing import Optional
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import aiohttp
import asyncio


"""
주요 변수
nowAction ( write / reply ) / 글쓰는지 댓글인지 체크
nowWriteStatus (optimize / basic ) / 최적화인지 아닌지 체크
allCount : 전체 변수, 글 / 댓글 나누기 위함
writeCount : 글쓰기 변수 nowAction 이 write일때 하나씩 증가, 최적화 아이디인지 판별하기 위함

cafe_id.cell(세로(열), 가로(행)).value
"""

###


def goScript(getDict):
    global driver
    
    pg.alert(text='시작 대기!!')
    
    
    
    
    

    

    cafe_optimize_file = load_workbook('./etc/naver_optimiz.xlsx')
    cafe_optimize = cafe_optimize_file.active
    cafe_id_file = load_workbook('./etc/naver_id.xlsx')
    cafe_id = cafe_id_file.active

    allCount = 0
    writeCount = 0
    endOptimize = ''
    nowAction = ''
    preIp = ''
    boardListKor = ['자유게시판', '수다방']
    boardListNum = [31, 34]

    while True:
        
        while True:
            getIP = changeIp()
            print(getIP)
            if not preIp == getIP:
                preIp = getIP
                break

        # 5로 나누어서 나머지가 1이면 (5의 배수 + 1 값이면 글쓰기 진행)
        allCount += 1
        nowActionNum = allCount % 5
        # nowActionNum = allCount % 2
        if nowActionNum == 1:
            nowAction = 'write'
            writeCount += 1
        else:
            nowAction = 'reply'
            nowWriteStatus = ""
        # 최적화 아이디인지 일반 아이디인지 체크
        
        chkOptimize1 = cafe_optimize.cell(1,2).value
        chkOptimize2 = cafe_optimize.cell(1,4).value
        if chkOptimize1 is None or chkOptimize2 is not None:
            endOptimize = 'on'
            
        if endOptimize == '' and nowAction == 'write':
            if os.path.exists(f'./etc/content/id_{writeCount}'):
                optimizeChkVal = cafe_optimize.cell(writeCount, 2).value
                if optimizeChkVal is None:
                    endOptimize = 'on'
                    nowWriteStatus = 'basic'
                else:
                    nowWriteStatus = 'optimize'
            else:
                endOptimize = 'on'
                nowWriteStatus = 'basic'
        else:
            nowWriteStatus = 'basic'

        # 최적화 글쓰기 / 일반 글쓰기 / 댓글쓰기 각 정보 (크롬정보 / 아이디 / 비번 / 게시판 번호 등) 부여하기
        if nowWriteStatus == 'optimize' and nowAction == 'write':
            # 최적화 아이디 일때
            uaSet = cafe_optimize.cell(writeCount, 1).value
            if uaSet is None:
                uaSet = getUaNum()
                cafe_optimize.cell(writeCount, 1).value = uaSet
                cafe_optimize_file.save('./etc/naver_optimiz.xlsx')

            nId = cafe_optimize.cell(writeCount, 2).value
            nPwd = cafe_optimize.cell(writeCount, 3).value
            nBoardName = cafe_optimize.cell(writeCount, 6).value
            nBoardNum = cafe_optimize.cell(writeCount, 7).value
        elif nowWriteStatus == 'basic' or nowAction == 'reply':
            # 일반 글쓰기 or 댓글 쓰기 일때 (안써진거 or 쓴지 3일 지난거)
            # 먼저 엑셀에서 사용한지 3일이 지난 값 가지고 오기
            nidExLength = getExLength(cafe_id)
            chkArr = asyncio.run(getEmptyArr(nidExLength, cafe_id))
            getRanVal = random.randrange(0, len(chkArr))
            getRanWorkVal = chkArr[getRanVal]

            uaSet = cafe_id.cell(getRanWorkVal, 1).value
            if uaSet is None:
                uaSet = getUaNum()
                cafe_id.cell(getRanWorkVal, 1).value = uaSet
                cafe_id_file.save('./etc/naver_id.xlsx')

            nId = cafe_id.cell(getRanWorkVal, 2).value
            nPwd = cafe_id.cell(getRanWorkVal, 3).value
            boardGetRan = random.randrange(0, 2)
            nBoardName = boardListKor[boardGetRan]
            nBoardNum = boardListNum[boardGetRan]

            print(datetime.now().date())
            cafe_id.cell(getRanWorkVal, 4).value = datetime.now().date()
            cafe_id_file.save('./etc/naver_id.xlsx')

        # 테스트겸 냅두자
        try:
            getVal = getRanWorkVal
        except:
            getVal = writeCount

        # pg.alert(text=f'{getVal}번째 있는 아이디로 {nowWriteStatus} {nowAction}작업, 크롬 정보 : {uaSet} / 아이디 : {nId} / 비번 : {nPwd} / 게시판 이름 {nBoardName}')
        print(f'{getVal}번째 있는 아이디로 {nowWriteStatus} {nowAction}작업, 크롬 정보 : {uaSet} / 아이디 : {nId} / 비번 : {nPwd} / 게시판 이름 {nBoardName}')

        if nowAction == 'write' and nowWriteStatus == 'basic':
            # 블로그 글따기 시작!!!
            # 엑셀로 랜덤 돌려서 제목 뽑기
            cafe_ex_file = load_workbook('./etc/subject_list.xlsx')
            cafe_ex = cafe_ex_file.active

            subjectCountArr = []
            for i in range(1, 5):
                k = 0
                while True:
                    k += 1
                    chkVal = cafe_ex.cell(k, i).value
                    if chkVal is None:
                        subjectCountArr.append(k)
                        break

            subjecArr = []
            for i, subjectCount in enumerate(subjectCountArr):
                if i == 1 or i == 2:
                    passNum = random.randrange(1, 3)
                    if passNum != 1:
                        continue
                getConNum = random.randrange(1, subjectCount)
                chkVal = cafe_ex.cell(getConNum, i+1).value
                subjecArr.append(chkVal)
            # 엑셀로 랜덤 돌려서 제목 뽑기 끝 이제 아래 블로그 컨텐츠 생성 함수에 넣고 막글 뽑자!

            blog_content = getBlogContent(subjecArr)
            subject = " ".join(subjecArr)
            with open("./etc/content/write_content.txt", "w") as f:
                f.write(subject)
                f.write('\n')
                f.write(blog_content)
            # 블로그 글따기 끝!!

        if nowAction == 'write':
            #네이버 메인에서 카페 진입 시작!
            with open(f'./etc/useragent/useragent_all.txt') as f:
                ua_data = f.readlines()[uaSet-1]

            options = Options()
            user_agent = ua_data
            options.add_argument('user-agent=' + user_agent)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(chrome_options=options, service=service)

            driver.get('https://www.naver.com')
            
            errchk = naverLogin(nId, nPwd)
            if errchk is not None:
                if nowWriteStatus == 'basic':
                    cafe_id.cell(getRanWorkVal, 4).value = errchk
                    cafe_id_file.save('./etc/naver_id.xlsx')
                elif nowWriteStatus == 'optimize':
                    cafe_optimize.cell(writeCount, 2).value = errchk
                    cafe_optimize_file.save('./etc/naver_optimiz.xlsx')
                allCount = allCount - 1
                driver.close()
                continue
            
            mainToCafe()

                
            writeBtn = searchElement('.inner_box .btn_write')
            untilEleGone(writeBtn[0], '.inner_box .btn_write')
            
            
            # 메뉴 선택하기
            selectBox = searchElement('.selectbox')
            untilEleShow(selectBox[0], '.layer_dimmed')
            selBoard = searchElement('.select_board li')
            for board in selBoard:
                if board.text == nBoardName:
                    board.click()
                    break
            
            # 메뉴 선택 완료! 글쓰기 시작~~~~
            if nowWriteStatus == 'basic':
                with open("./etc/content/write_content.txt", "r") as f:
                    getContents = f.readlines()
            elif nowWriteStatus == 'optimize':
                with open(f'./etc/content/id_{writeCount}/content.txt', 'r') as f:
                    getContents = f.readlines()
            
            chkImg = [0,0,0,0,0,0]
            
            # 글 작성 전 창 체크
            focus_window('글쓰기')
            
            articleWriteFormSubject = searchElement('.ArticleWriteFormSubject')
            articleWriteFormSubject[0].click()
            keyboard.write(text=getContents[0], delay=0.3)
            
            oneEditor = searchElement('#one-editor')
            oneEditor[0].click()
            for i, line in enumerate(getContents):
                if i == 0:
                    continue
                try:
                    line_temp = line.replace('\n','')
                    chkImg = line_temp.split('|')
                except:
                    pass
                
                if chkImg[0] == 'img_line':
                    imageUpload = searchElement('.se-toolbar-item-image')
                    imageUpload[0].click()
                    
                    nowPath = os.getcwd()
                    if nowWriteStatus == 'basic':
                        imagePath = nowPath + "\etc\content\images"
                        imageList = os.listdir(imagePath)
                        getImage = imageList[random.randrange(0, len(imageList))]
                        wait_float(1.5, 2.2)
                        pyperclip.copy(imagePath)
                        wait_float(0.5, 0.9)
                        pg.hotkey('ctrl','v')
                        wait_float(0.5, 0.9)
                        pg.press('enter')
                        
                        wait_float(0.9, 1.6)
                        pyperclip.copy(getImage)
                        wait_float(0.5, 0.9)
                        pg.hotkey('ctrl','v')
                        wait_float(0.5, 0.9)
                        pg.press('enter')
                        
                        wait_float(3.5, 4.8)
                        
                    elif nowWriteStatus == 'optimize':
                        imagePath = nowPath + f"\etc\content\id_{writeCount}"
                        wait_float(1.5, 2.2)
                        pyperclip.copy(imagePath)
                        wait_float(0.5, 0.9)
                        pg.hotkey('ctrl','v')
                        wait_float(0.5, 0.9)
                        pg.press('enter')
                        
                        wait_float(0.9, 1.6)
                        pyperclip.copy(chkImg[1])
                        wait_float(0.5, 0.9)
                        pg.hotkey('ctrl','v')
                        wait_float(0.5, 0.9)
                        pg.press('enter')
                        
                        wait_float(3.5, 4.8)
                    # 끝난다음 초기화
                    chkImg = [0,0,0,0,0,0]
                elif line == 'enter':
                    pg.press('enter')
                else:
                    keyboard.write(text=line, delay=0.1)
            successBtn = searchElement('.GnbBntRight__green')
            untilEleGone(successBtn[0], '.GnbBntRight__green')
            
            
            getLinkMore = searchElement('.btn_aside .more')
            getLinkMore[0].click()
            
            getLink = searchElement('.layer_list li')
            getLink[3].click()
            
            getLinkData = cb.paste()
            if nowWriteStatus == 'optimize':
                with open(f'./etc/content/id_{writeCount}/reply.txt', 'r') as f:
                    getTempReplys = f.readlines()
                    getTempReplys.insert(0, '0\n')
                    getTempReplysName_temp = getLinkData.split('/')
                    getTempReplysName = getTempReplysName_temp[-1]
                with open(f'./etc/content/temp_reply/{getTempReplysName}.txt', 'w') as f:
                    f.writelines(''.join(getTempReplys))
            
            with open('./etc/work_link.txt', 'a') as f:
                f.write('\n')
                f.write(getLinkData)

            
            with open('./etc/work_link.txt', 'r') as f:
                chkLines = f.readlines()

            if len(chkLines) > 5:
                getDelOptimizeReplyNum_temp = chkLines[0].split('/')
                getDelOptimizeReplyNum = getDelOptimizeReplyNum_temp[-1]
                if os.path.exists(f'./etc/content/temp_reply/{getDelOptimizeReplyNum}.txt'):
                    os.remove(f'./etc/content/temp_reply/{getDelOptimizeReplyNum}.txt')
                del chkLines[0]
                
            with open('./etc/work_link.txt','w') as f:
                f.writelines(''.join(chkLines))
            
       
        # ★★★★★★★★ 댓글 작성 시작!!
        if nowAction == 'reply':
            with open(f'./etc/useragent/useragent_all.txt') as f:
                ua_data = f.readlines()[uaSet-1]

            options = Options()
            user_agent = ua_data
            options.add_argument('user-agent=' + user_agent)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(chrome_options=options, service=service)

            driver.get('https://www.naver.com')
            
            errchk = naverLogin(nId, nPwd)
            if errchk is not None:
                if nowWriteStatus == 'basic':
                    cafe_id.cell(getRanWorkVal, 4).value = errchk
                    cafe_id_file.save('./etc/naver_id.xlsx')
                elif nowWriteStatus == 'optimize':
                    cafe_optimize.cell(writeCount, 2).value = errchk
                    cafe_optimize_file.save('./etc/naver_optimiz.xlsx')
                allCount = allCount - 1
                driver.close()
                continue
            
            mainToCafe()
        else:
            goToHome = searchElement('.header h1')
            untilEleGone(goToHome[0], '.post_title')
        
        
        # 카페 메인 진입 끝! 게시글 클릭 시작!
        nowWriteStatus = ''
        with open(f'./etc/work_link.txt') as f:
            workLinkList = f.readlines()
        
        for workLink in workLinkList:
            workLink_temp = workLink.replace('\n', '')
            workLinkOn = workLink_temp.split('/')[-1]
            boardListAll = searchElement('.list_area li')
            for boardList in boardListAll:
                chkBoardLink = boardList.find_element(by=By.CSS_SELECTOR, value='.txt_area').get_attribute('href')
                chkVal = workLinkOn in chkBoardLink
                
                if chkVal:
                    untilEleGone(boardList, '.txt_area')
                    pg.moveTo(300, 500)
                    randomFor = random.randrange(2,5)
                    for i in range(1, randomFor):
                        wait_float(1.5, 2.5)
                        pg.scroll(-500)
                    break
            
            # 게시글 클릭 완료 댓글 쓰기 시작!
            
            randomActionVal = random.randrange(1, 4)
            print(randomActionVal)
            if randomActionVal == 1:
                
                # 댓글 작성 전 창 체크
                replyGoBtn = searchElement('.f_reply')
                untilEleShow(replyGoBtn[0], '.HeaderIcon')
                
                wait_float(0.5, 0.9)
                replyBtn = searchElement('.comment_textarea')
                replyBtn[0].click()
            
                getReply = ""
                print(workLinkOn)
                if os.path.exists(f'./etc/content/temp_reply/{workLinkOn}.txt'):
                    
                    with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'r') as f:
                        getTempReplys = f.readlines()
                        getTempNum = getTempReplys[0].replace('\n', '')
                        try:
                            getReply = getTempReplys[int(getTempNum) + 1]
                            getTempReplys[0] = str(int(getTempNum) + 1) + '\n'
                        except:
                            getReply = ""
                    wait_float(0.2,0.7)
                    if getReply != "":
                        with open(f'./etc/content/temp_reply/{workLinkOn}.txt', 'w') as f:
                            f.writelines(''.join(getTempReplys))
                wait_float(0.2,0.7) 
                if getReply == "":
                    with open(f'./etc/all_reply.txt', 'r') as f:
                        getTempReplysAll = f.readlines()
                        getTempRanNum = random.randrange(0, len(getTempReplysAll))
                        getReply = getTempReplysAll[getTempRanNum]
                        
                focus_window('카페')
                keyboard.write(text=getReply, delay=0.1)
                replySuccessBtn = searchElement('.btn_done')
                untilEleGone(replySuccessBtn[0], '.btn_done')
                
                goToPostiongBtn = searchElement('.HeaderGnbLeft')
                untilEleGone(goToPostiongBtn[0], '.HeaderGnbLeft')
                
            goToHome = searchElement('.header h1')
            untilEleGone(goToHome[0], '.post_title')
                
        driver.close()
                     
                
                
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염

# 상품 들어가서 스크롤 내리고 나오기




# async def wait헤arr, ex, i, chk):
#     target_click = int(ex.cell(i, 4).value)
#     now_click = ex.cell(i, 5).value

#     if now_click is None:
#         ex.cell(i, 5).value = 0
#         now_click = 0
#     now_click = int(now_click)
#     if chk == 'Y':
#         if now_click < target_click:
#             arr.append(i)
#     else:
#         if now_click >= target_click:
#             arr.append(i)


# async def playAsync_getArr(arr, ex, linkCount, chk):
#     try:
#         await asyncio.gather(*[waitPrint(arr, ex, i, chk) for i in range(1, linkCount + 1)])
#     except:
#         pass


# 결과 값(workarr) 을 가지고 해당 인덱스의 엑셀에 1씩 더하기
# async def linkExcelPlus(ex, val):
#     setVal = ex.cell(val, 5).value
#     ex.cell(val, 5).value = setVal + 1


# async def playAsync_plusArr(arr, ex):
#     try:
#         await asyncio.gather(*[linkExcelPlus(ex, val) for val in arr])
#     except:
#         pass


def naverLogin(load_id, load_pass):
    search_bar = searchElement(".sch_ico_aside")
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap")
    login_btn[0].click()

    # 로그인 부분
    
    focus_window('로그인')
    while True:
        searchElement("#id")

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header")

        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close")
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp


def searchElement(ele):
    wait_float(0.3, 0.7)
    re_count = 1
    element = ""
    while True:
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            re_count += 1

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
            try:
                btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
                if btnEle is not None:
                    return
            except:
                continue
        except:
            pass


def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
            try:
                btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
                if btnEle is None:
                    return
            except:
                return
        except:
            pass


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent




async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def mainToCafe():
    shs_item = searchElement('.shs_item')
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break
    
    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe')



    myCafeList = searchElement('.list_cafe__favorites li')
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")
    
    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break
        
    # 카페 진입 끝

def getBlogContent(subjectArr):
    # 블로그 글따기 시작

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get('https://section.blog.naver.com/BlogHome.naver')

    try:
        popup = driver.find_element(
            by=By.CSS_SELECTOR, value='#floatingda_home')
        popupClostBtn = popup.find_elements(by=By.CSS_SELECTOR, value='button')
        popupClostBtn[-1].click()
    except:
        pass

    while True:
        nCategoryList = driver.find_elements(
            by=By.CSS_SELECTOR, value='.navigator_category a')
        categoryRanVal = random.randrange(1, len(nCategoryList))
        nCategoryList[categoryRanVal].click()

        wait_float(0.5, 0.9)

        paginationNum = driver.find_elements(
            by=By.CSS_SELECTOR, value='.pagination span')
        driver.execute_script(
            "arguments[0].scrollIntoView();", paginationNum[0])
        paginationRanVal = random.randrange(1, len(paginationNum))
        getClickPage = paginationNum[paginationRanVal].find_element(
            by=By.CSS_SELECTOR, value='a')
        getClickPage.click()
        wait_float(0.5, 0.9)

        infoPostList = driver.find_elements(
            by=By.CSS_SELECTOR, value='.info_post')
        infoPostRanVal = random.randrange(1, len(infoPostList))
        getInfoPostTag_a = infoPostList[infoPostRanVal].find_element(
            by=By.CSS_SELECTOR, value='.desc a')
        getInfoPostLink = getInfoPostTag_a.get_attribute('href')
        getInfoPostLink = getInfoPostLink.replace('//', '//m.')

        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')

        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        allStr = "".join(allStr)
        if len(allStr) > 600 and len(allStr) < 1200:
            len(allStr)
            break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    driver.close()
    return allContent


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
