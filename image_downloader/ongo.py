
import random
import threading
import time
from datetime import date, datetime, timedelta
from dateutil.relativedelta import *
import sys
import os
from pathlib import Path
from typing import Optional
from pyparsing import And
import requests
from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import aiohttp
import asyncio
from PIL import Image



###


def goScript(getDict):
    pass

def imageReduceCapa():
    # 1. 이미지 파일 기본 정보 읽기 #
    print('시작~~~~~~~~~~~~~')
    try:
        im = Image.open("//pre_image//anchors.png")
        img_width, img_height = im.size
        print("이미지 확장자:", im.format)
        print("이미지 사이즈:", im.size)
        print("이미지 가로:", img_width)
        print("이미지 세로:", img_height)
        print("이미지 모드:", im.mode)
    except OSError as e:
        print('에러입니당~~~~~~~~~~~~~~~~~~')
        print(e)


def naverLogin(load_id, load_pass):
    search_bar = searchElement(".sch_ico_aside")
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap")
    login_btn[0].click()

    # 로그인 부분

    focus_window('로그인')
    while True:
        searchElement("#id")

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header")

        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass
        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".action_inner")
            if unPwd:
                return "비정상적 활동"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close")
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp


def changeIpSpeed():
    pg.alert(driver)
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    while True:
        try:

            ondevice.shell("input keyevent KEYCODE_POWER")
            ondevice.shell("svc data disable")
            ondevice.shell("settings put global airplane_mode_on 1")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

            ondevice.shell("svc data enable")
            ondevice.shell("settings put global airplane_mode_on 0")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
            time.sleep(3)
            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
        except:

            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue

        driver.get('https://fast.com/ko/')
        searchElement('.speed-results-container')
        time.sleep(3)
        getInternetRapidEle = searchElement('.speed-results-container')
        getInternetRapid = getInternetRapidEle[0].text
        if float(getInternetRapid) < 2.7:
            continue
        else:
            driver.close()
            break

    return getIp
# def searchElement(ele):
#     wait_float(0.3, 0.7)
#     re_count = 0
#     element = ""
#     while True:
#         re_count += 1
#         if re_count % 5 == 0:
#             print(ele)
#             print("새로고침!!!!")
#             driver.refresh()
#             focus_window('chrome')
#             pg.press('F5')

#             wait_float(2.5, 3.7)
#         try:
#             driver.implicitly_wait(3)
#             selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
#             if selected_element:
#                 wait_float(2.5, 3.7)
#                 return selected_element
#             # element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
#         except:
#             pass


def searchElement(ele):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


# def untilEleShow(clickEle, searchEle):
#     while True:
#         try:
#             clickEle.click()
#             time.sleep(1)
#             try:
#                 btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
#                 if btnEle is not None:
#                     return
#             except:
#                 continue
#         except:
#             pass

def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


# def untilEleGone(clickEle, searchEle):
#     while True:
#         try:
#             clickEle.click()
#             time.sleep(1)
#             try:
#                 btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
#                 if btnEle is None:
#                     return
#             except:
#                 return
#         except:
#             pass

def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass

        try:
            btnEle = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent


async def getEmptyArr(setNum, exName):
    getArr = []
    asyncio.gather(*[busyFunc(i, getArr, exName) for i in range(1, setNum)])
    return getArr


async def busyFunc(i, getArr, exName):
    getTime = exName.cell(i, 4).value

    try:
        if getTime is None:
            getTime = datetime.now().date()
        if isinstance(getTime, datetime):
            getTime = getTime.date()
        compareTime = datetime.now() - timedelta(days=3)
        if exName.cell(i, 4).value is None or getTime <= compareTime.date():
            getArr.append(i)
    except:
        pass


def getExLength(exName):
    ExLength = 0
    while True:
        ExLength += 1
        if exName.cell(ExLength, 2).value is None:
            break
    return ExLength


def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet


def mainToCafe():
    shs_item = searchElement('.shs_item')
    for item in shs_item:
        chkCafe = item.find_element(
            by=By.CSS_SELECTOR, value='a').get_attribute('href')
        if 'cafe' in chkCafe:
            untilEleGone(item, '.shs_list')
            break

    myCafeGo = searchElement('.mycafe .btn_cafe_more')
    untilEleGone(myCafeGo[0], '.mycafe')

    myCafeList = searchElement('.list_cafe__favorites li')
    with open("./etc/cafe_info.txt", "r") as f:
        getCafeNameList = f.readlines()
        getCafeName = getCafeNameList[0]
        getCafeName = getCafeName.replace(" ", "")

    for onCafe in myCafeList:
        chkCafeTitle = onCafe.find_element(
            by=By.CSS_SELECTOR, value='.title').text
        chkCafeTitle = chkCafeTitle.replace(" ", "")

        if chkCafeTitle in getCafeName:
            untilEleGone(onCafe, '.list_cafe__favorites')
            break

    # 카페 진입 끝


def getBlogContentChrome(subjectArr):
    pg.alert('함수 진입 대기~~~')
    
    with open('./etc/find_keyword.txt', 'r') as r:
        allKeyword = r.readlines()
    
    keyCount = random.randrange(0, len(allKeyword))
    getKeyword = allKeyword[keyCount]
    getKeyword = getKeyword.replace('\n', '')
    print(getKeyword)
    driver.get('https://www.google.com/')
    
    
    searchBar = searchElement('.gLFyf')
    wait_float(0.5,0.8)
    searchBar[0].click()
    wait_float(0.5,0.8)
    searchBar[0].send_keys(f'site:blog.naver.com {getKeyword}')
    pg.press('enter')
    wait_float(0.8,1.5)
    pg.press('enter')
    
    
    
    getTools = searchElement('.t2vtad')
    getTools[0].click()
    wait_float(0.8,1.5)
    
    getToolsIf = searchElement('.KTBKoe')
    for getToolsIfOn in getToolsIf:
        if '날짜' in getToolsIfOn.text:
            getToolsIfOn.click()
    wait_float(0.8,1.5)
            
    getPeriodIf = searchElement('.y0fQ9c')
    for getPeriodIfOn in getPeriodIf:
        if '설정' in getPeriodIfOn.text:
            getPeriodIfOn.click()
    wait_float(0.8,1.5)
    
    today = datetime.today()
    print(today)

    this_month_first = datetime(today.year, today.month, 1)
    chkMonthFirst = this_month_first.strftime('%m/%d/%Y')
    driver.find_element(by=By.CSS_SELECTOR, value='.OouJcb').send_keys(chkMonthFirst)
    
    
    next_month = datetime(today.year, today.month, 1) + relativedelta(months=1)
    this_month_last = next_month + relativedelta(seconds=-1)
    chkMonthLast = this_month_last.strftime('%m/%d/%Y')
    driver.find_element(by=By.CSS_SELECTOR, value='.rzG2be').send_keys(chkMonthLast)
    wait_float(0.8,1.5)
    pg.press('enter')
    
    getPagingList = searchElement('.NVbCr')
    getPgCount = random.randrange(2,len(getPagingList) - 1)
    getPagingList[getPgCount].click()
    
    pg.alert('대기~~~')
    
    
    
    # #시작일
    # OouJcb
    
    # #종료일
    # rzG2be
    
    
    
    
    
    
    
    
    
    
    
     
    

def getBlogContent(subjectArr):
    # 블로그 글따기 시작

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get('https://section.blog.naver.com/BlogHome.naver')

    try:
        popup = driver.find_element(
            by=By.CSS_SELECTOR, value='#floatingda_home')
        popupClostBtn = popup.find_elements(by=By.CSS_SELECTOR, value='button')
        popupClostBtn[-1].click()
    except:
        pass

    while True:

        try:
            nCategoryList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.navigator_category a')
            categoryRanVal = random.randrange(0, len(nCategoryList) - 1)
            nCategoryList[categoryRanVal].click()

            wait_float(0.5, 0.9)

            paginationNum = driver.find_elements(
                by=By.CSS_SELECTOR, value='.pagination span')
            driver.execute_script(
                "arguments[0].scrollIntoView();", paginationNum[0])
            paginationRanVal = random.randrange(0, len(paginationNum) - 1)
            getClickPage = paginationNum[paginationRanVal].find_element(
                by=By.CSS_SELECTOR, value='a')
            getClickPage.click()
            wait_float(0.5, 0.9)

            infoPostList = driver.find_elements(
                by=By.CSS_SELECTOR, value='.info_post')
            infoPostRanVal = random.randrange(0, len(infoPostList) - 1)
            getInfoPostTag_a = infoPostList[infoPostRanVal].find_element(
                by=By.CSS_SELECTOR, value='.desc a')
            getInfoPostLink = getInfoPostTag_a.get_attribute('href')
            getInfoPostLink = getInfoPostLink.replace('//', '//m.')
        except:
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
            wait_float(2.5, 3.5)
            continue

        page = requests.get(getInfoPostLink)
        soup = bs(page.text, "html.parser")
        elements = soup.select('.se-module.se-module-text')

        allStr = []
        for ele in elements:
            p = re.compile('[\uAC00-\uD7A30-9a-zA-Z\s]+')
            chkResult = p.findall(str(ele))
            allStr = allStr + chkResult

        p_str = re.compile(r'[a-zA-Z0-9,|\n]+')
        p_space = re.compile('\s\s')

        for i in range(1, len(allStr)):
            for j, strin in enumerate(allStr):
                getStr = p_str.search(strin)
                if getStr is not None:
                    allStr.pop(j)
                    break
                getSpace = p_space.search(strin)
                if getSpace is not None:
                    allStr.pop(j)
                    break
                if strin == " ":
                    allStr.pop(j)
                    break
        allStr = "".join(allStr)
        if len(allStr) < 600:
            continue
        if len(allStr) > 1200:
            sliceRanNum = random.randrange(1050, 1150)
            allStr = allStr[0:sliceRanNum]
        break

    resetStrArr = allStr.split(' ')

    resetListArr = list_chunk(resetStrArr, 12)
    for resetList in resetListArr:
        setRan = random.randrange(2, 5)
        resetOn = random.sample(range(1, 13), setRan)

        if resetList == "":
            continue

        for inon in resetOn:
            changeRanCount = random.randrange(0, len(subjectArr))
            chkChangeRan = random.randrange(1, 6)
            if chkChangeRan == 1:
                try:
                    resetList[inon - 1] = subjectArr[changeRanCount]
                except:
                    pass
            else:
                try:
                    resetList[inon - 1] = ''
                except:
                    pass

    imgLineCountBasic = divmod(len(resetListArr), 2)
    imgLineCount = random.randrange(
        int(imgLineCountBasic[0]) - 4, int(imgLineCountBasic[0]) + 4)

    allContent = ''
    for i, setList in enumerate(resetListArr):
        if imgLineCount == i:
            allContent = allContent + 'img_line|randomimg\n'
        for setStr in setList:
            if setStr == '':
                continue
            elif len(setStr) > 20:
                continue
            allContent = allContent + setStr
            allContent = allContent + ' '
        allContent = allContent + '\n'

    driver.close()
    return allContent


# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]
