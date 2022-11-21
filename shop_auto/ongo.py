import linecache
import random
import numpy as np
import threading
import time
import sys
import os
from pathlib import Path
from typing import Optional
import requests
import json
import re

import pyautogui as pg
import pyperclip
import pywinauto
import pygetwindow as gw
import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import requests
import winsound as ws
import glob
import random
import aiohttp
import asyncio


def goScript(getDict):
    
    global driver
    
    if getDict['backVal'] == 1:
        backRanVal = random.randrange(0,2)
        backRanVal = 0
        if backRanVal == 1:
            with open('./etc/useragent/useragent_all.txt','r') as r:
                uaAll = r.readlines()
            rVal = random.randrange(0,len(uaAll))
            getUa = uaAll[rVal]
            
            options = Options()
            options.add_argument('user-agent=' + getUa)
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(chrome_options=options, service=service)
            
            
            
            driver.get('https://www.naver.com')
            
            backWb = load_workbook('./etc/backlinks.xlsx')
            backEx = backWb.active
            searchKeyword = backEx.cell(1,1).value
            
            while True:
                wait_float(0.3,0.9)
                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='#MM_SEARCH_FAKE')
                    break
                except:
                    pass
                
                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='#query')
                    break
                except:
                    pass
                
                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='#nx_query')
                    break
                except:
                    pass
            searchBar.click()
            wait_float(1.5,2.3)
            keyboard.write(text=searchKeyword, delay=0.3)
            wait_float(0.3,0.9)
            pg.press('enter')
            searchElement('.total_wrap')
            clickEle = searchElement(backEx.cell(1,3).value)
            clickEle[0].click()
            
            wait_float(2.5, 4.5)
            pg.moveTo(200, 200)
            maxRange = random.randrange(3,6)
            forCount = 0
            while maxRange > forCount:
                scrollVal = random.randrange(300, 500)
                pg.scroll(-scrollVal)
                wait_float(2.5, 4.5)
                forCount += 1
            driver.back()
            
            
            
            
            
            
            
            
    
    # 전체 반복 시작 전 preIp 값 초기화
    preIp = ""

    # 전체 반복 시작 전 전체 카운트값 초기화 (작업할 즉, 비어있는)
    idp_wb = load_workbook('./etc/naver_id.xlsx')
    id_excel = idp_wb.active
    allCount = 0
    chk_login_val = "wait"
    while chk_login_val != None:
        allCount += 1
        chk_login_val = id_excel.cell(allCount, 4).value
    allCount = allCount

    # 전체 반복 시작 전 지쇼 링크 열고 전체 행 갯수 체크
    jisho_wb = load_workbook('./etc/jisho_link.xlsx')
    link_excel = jisho_wb.active
    
    
    setVal = "wait"
    linkCount = 1
    while setVal != None:
        linkCount += 1
        setVal = link_excel.cell(linkCount, 2).value
    linkCount = linkCount - 1

    while True:
        
        startTime = time.time()
        # 1/3 확률로 네이버 아이디 비번 저장 (로그인 준비)

        if getDict['loginVal'] == 1:
            chk_login = random.randrange(1, 8)
        else:
            chk_login = 1
            load_id = ""
        
        if chk_login != 1:
            while True:
                wait_float(0.2, 0.5)
                load_connect_info = id_excel.cell(allCount, 1).value
                if load_connect_info is None:
                    load_connect_info = getUaNum()
                    link_excel.cell(allCount, 1).value = load_connect_info
                    link_excel.save('./etc/naver_id.xlsx')
                    
                    
                load_id = id_excel.cell(allCount, 2).value
                load_pass = id_excel.cell(allCount, 3).value

                if load_connect_info != '' and load_id != '' and load_pass != '':
                    break

        # 로그인 준비 끝

        # 아이피 체크 (기존 아이피와 같으면 다시, 아니면 break)
        if getDict['ipval'] == 1:
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service)
            while True:
                getIP = changeIpSpeed()
                print(getIP)
                if not preIp == getIP:
                    preIp = getIP
                    break
        
        # 아이피 체크 끝

        """
        작업 할 지식쇼핑 링크 구하기
        (총 링크가 5개 이하면 전부다, 6~10 개면 4개, 10개 이상이면 8개씩)
        1. 전체 항목 목표 클릭수와 현재 클릭수 비교 (비동기 항목으로 보냄)
        2. 현재 클릭수가 목표 클릭수보다 크거나 같으면 일단 탈락
        3. 남은것들이 총 링크와 같으면 위 조건으로, 총 링크보다 작은데 미달이면 미달인 갯수만큼 아무거나 채우기, 총 링크보다 크면 그만큼 빼기
        4. 그래서 해서 행 번호만 뽑아서 배열로 저장
        """

        workArr = []
        tempWorkArr = []
        if linkCount <= 5:
            searchCount = linkCount
        elif linkCount >= 6 and linkCount < 10:
            searchCount = 6
        else:
            searchCount = 9
        asyncio.run(playAsync_getArr(workArr, link_excel, linkCount, 'Y'))
        random.shuffle(workArr)

        if len(workArr) > searchCount:
            workArr = workArr[0:searchCount]
        else:
            asyncio.run(playAsync_getArr(tempWorkArr, link_excel, linkCount, 'N'))
            random.shuffle(tempWorkArr)
            getCountLen = searchCount - len(workArr)
            tempWorkArr = tempWorkArr[0:getCountLen]
            workArr = workArr + tempWorkArr
            # workArr = workArr[0:searchCount]
            random.shuffle(workArr)
            
            
            

        asyncio.run(playAsync_plusArr(workArr, link_excel))
        jisho_wb.save('./etc/jisho_link.xlsx')
        
        # 작업할 배열 구하기 끝~~~~~

        # 접속할 USER AGENT 설정
        if chk_login != 1:
            # connect_info = load_connect_info.split(',')
            # with open(f'./etc/useragent/useragent_{connect_info[0]}.txt') as f:
            #     ua_data = f.readlines()[int(connect_info[1]) - 1]
            with open(f'./etc/useragent/useragent_all.txt') as f:
                ua_data = f.readlines()[load_connect_info]
                ua_data = ua_data.replace('\n', '')
                
        else:
            ua_data = linecache.getline(
                './etc/useragent/useragent_all.txt', random.randrange(1, 14)).strip()
        # 설정 끝~ 접속하기

        options = Options()
        user_agent = ua_data
        options.add_argument('user-agent=' + user_agent)

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(chrome_options=options, service=service)

        driver.get('https://www.naver.com')

        time.sleep(2)
        focus_window('NAVER')

        # 네이버 로그인
        if chk_login != 1:
            errchk = naverLogin(load_id, load_pass)
            if errchk is not None:
                id_excel.cell(allCount, 4).value = errchk
                idp_wb.save('./etc/naver_id.xlsx')
                allCount += 1
                continue
            main_menus = searchElement(".shs_link")
            for main_menu in main_menus:
                if 'mail' in main_menu.get_attribute('href'):
                    main_menu.click()
                    break
            wait_float(2.5, 3.5)
            
            while True:
                wait_float(0.3, 0.5)
                try:
                    driver.find_element(by=By.CSS_SELECTOR, value='#MM_SEARCH_FAKE')
                    break
                except:
                    driver.back()
        mainSearch = searchElement("#MM_SEARCH_FAKE")
        mainSearch[0].click()

        subSearch = searchElement("#query")

        focus_window('NAVER')

        # subSearch[0].send_keys('네이버쇼핑')
        subSearch[0].click()
        keyboard.write(text="네이버쇼핑", delay=0.3)

        searchSubmit = searchElement(".MM_SEARCH_SUBMIT")
        searchSubmit[0].click()

        nShoppingLink = searchElement(".link_name")

        # nShoppingLink[0].click()
        untilEleGone(nShoppingLink[0], ".link_name")

        for workVal in workArr:
            
            searchKeyword = link_excel.cell(workVal, 2).value
            searchJisho(searchKeyword, driver)

            nShopCategory = searchElement(".mainFilter_option__c4_Lq")
            
            try:
                UseLessBtn = driver.find_element(by=By.CSS_SELECTOR, value='.basicFilter_btn_close__qftDk svg')
                UseLessBtn.click()
            except:
                pass
            
            

            setTong = link_excel.cell(workVal, 1).value
            if setTong is not None:
                chkin_tong = ""
                for category in nShopCategory:
                    category_name = category.text.replace('+', '')
                    if category_name[0:2] == setTong:
                        chkin_tong = "on"
                        print('여기서 나는 에러가 맞는걸까요?????')
                        # untilEleShow(category, ".selected_btn_del__0mIMB")
                        untilEleShow(category, ".mainFilter_option__c4_Lq")

                if chkin_tong == "":
                    addKeyword = link_excel.cell(workVal, 1).value
                    searchRan = random.randrange(0, 2)
                    if searchRan == 0:
                        searchKeyword = searchKeyword + " " + addKeyword
                    else:
                        searchKeyword = addKeyword + " " + searchKeyword
                    searchJisho(searchKeyword, driver)
                    
            # 상위 4개 중 1개 클릭
            
            # 여기서 6개까지 찾고 / 그중에 있으면 그냥 한번만, 없으면 원래대로
            
            highWork = ""
            item_list = driver.find_elements("xpath", "//*[contains(@class, 'product_list_item')]")
            print('상위 작업 체크 시작!!')
            chkCount = 0
            for highCount in range(6):
                chkCount += 1
                getHighHref = item_list[highCount].find_element(by=By.CSS_SELECTOR, value='a').get_attribute('href')
                searchMid = link_excel.cell(workVal, 3).value
                if str(searchMid) in getHighHref:
                    highWork = "on"
                    driver.execute_script("arguments[0].scrollIntoView();", item_list[highCount])
                    item_list[highCount].click()
                    maxRange = random.randrange(7, 10)
                    onProductScroll(maxRange)
                    
                    #찜하기~~~
                    zzimRandomVal = random.randrange(1, 4)
                    if chk_login != 1 and zzimRandomVal == 1:
                        zzimAction(chkCount, workVal, link_excel, jisho_wb)
                        
                    break
            
            print('상위 작업 체크 끝~~~~!!')
                
            # 상위에 있는거 찾는거 끝
            if highWork == "":
                item_list = driver.find_elements("xpath", "//*[contains(@class, 'product_list_item')]")
                topProduct_val = random.randrange(0, 4)
                wait_float(0.5, 1.7)
                driver.execute_script("arguments[0].scrollIntoView();", item_list[topProduct_val])
                untilEleGone(item_list[topProduct_val], ".product_list_item")

                wait_float(2, 5)

                maxRange = random.randrange(2, 4)
                onProductScroll(maxRange)

                truncBreak = ""
                truncCount = 1
                while True:
                    truncCount += 1
                    
                    resetCount = 0
                    while True:
                        resetCount += 1
                        if resetCount > 20:
                            driver.refresh()
                            wait_float(2, 4)
                            resetCount = 0
                        
                        item_list = driver.find_elements("xpath", "//*[contains(@class, 'product_list_item')]")

                        if len(item_list) < 35:
                            pg.hotkey('end')
                            wait_float(2, 4)
                        else:
                            break
                        
                    chkCount = 0
                    for item in item_list:
                        chkCount += 1
                        getHref = item.find_element(by=By.CSS_SELECTOR, value='a').get_attribute('href')
                        searchMid = link_excel.cell(workVal, 3).value
                        wait_float(0.1, 0.3)
                        if str(searchMid) in getHref:
                            truncBreak = "on"
                            # action.move_to_element(item).perform()
                            driver.execute_script("arguments[0].scrollIntoView();", item)
                            item.click()
                            maxRange = random.randrange(4, 6)
                            onProductScroll(maxRange)
                            break
                        
                        
                        
                    #찜하기~~~
                    zzimRandomVal = random.randrange(1, 4)
                    if chk_login != 1 and zzimRandomVal == 1:
                        zzimAction(chkCount, workVal, link_excel, jisho_wb)
                        
                        
                                

                    if truncBreak == "on":
                        break

                    pageBtn = driver.find_elements(by=By.CSS_SELECTOR, value='.paginator_list_paging__VxWMC > a')
                    for btn in pageBtn:
                        if int(btn.text) == truncCount:
                            btn.click()
                            break

        # 끝내고 allCount 값 ++
        driver.quit()
        if chk_login != 1:
            id_excel.cell(allCount, 4).value = "지식쇼핑 작업 완료"
            idp_wb.save('./etc/naver_id.xlsx')
            allCount += 1
            
        # 아래 내용 웹훅 넣기
        endTime = time.time() - startTime
        
        webhook_url = "https://adpeak.kr/chk_jisho/"
        data = {'on_time' : endTime}
        requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)
        r = requests.post(webhook_url, data=json.dumps(data), headers={'Content-Type' : 'application/json'}, verify=False)
        
        
        
        
        
        
        
        
def zzimAction(chkCount, workVal, link_excel, jisho_wb):
    while True:
        try:
            allItem = driver.find_elements("xpath", "//*[contains(@class, 'product_list_item')]")
            if allItem:
                break
        except:
            continue
        
    get_item = allItem[chkCount - 1]
    
    try:
        itemZzimEle = get_item.find_elements(by=By.CSS_SELECTOR, value='.product_info_count__PSSO1 span')
        itemZzimLastEle = itemZzimEle[-1]
        itemZzimCountText = itemZzimLastEle.text
        itemZzimCount = re.sub(r'[^0-9]', '', itemZzimCountText)
        itemZzimCount = int(itemZzimCount)
    except:
        itemZzimCount = 0
    
    targetZzimCountChk = link_excel.cell(workVal, 6).value
    if targetZzimCountChk is not None:
        targetZzimCount = int(targetZzimCountChk)
    else:
        targetZzimCount = 0
        
    if targetZzimCount > itemZzimCount:
        itemZzim = get_item.find_element(by=By.CSS_SELECTOR, value='.product_btn_zzim__kfwDI')
        print(itemZzim)
        if itemZzim.text == '찜하기':
            writeZzim = itemZzimCount + 1
            link_excel.cell(workVal, 7).value = writeZzim
            jisho_wb.save('./etc/jisho_link.xlsx')
            itemZzim.click()
        
        
        
        


def ongo_searchItem():
    URL = "https://openapi.naver.com/v1/search/shop"
    headers = {"X-Naver-Client-Id": get_secret(
        'NAVER_API_ID'), "X-Naver-Client-Secret": get_secret('NAVER_API_SECRET')}
    # 전체 반복 시작 전 지쇼 링크 열고 전체 행 갯수 체크
    jisho_wb = load_workbook('./etc/jisho_link.xlsx')
    link_excel = jisho_wb.active
    setVal = "wait"
    linkCount = 1

    while setVal != None:
        linkCount += 1
        setVal = link_excel.cell(linkCount, 2).value

    print(linkCount)

    for i in range(1, linkCount):
        keyword = link_excel.cell(i, 2).value
        tong = link_excel.cell(i, 1).value
        productId = link_excel.cell(i, 3).value

        keyword = keyword.strip()

        if tong is not None:
            tong = tong.strip()
            if tong == "SK":
                tong = "SKT"
            elif tong == "LG":
                tong = "LG U+"

        productId = str(productId)

        allCount = 0
        itemCount = 0
        chk_loop = ""
        while chk_loop == "":
            try:
                params = {'query': keyword, 'start': allCount *
                          100 + 1, 'display': '100'}
                res = requests.get(URL, headers=headers, params=params).json()
                for item in res['items']:
                    if tong is not None:
                        if item['category3'] == tong:
                            itemCount += 1
                    else:
                        itemCount += 1
                    if item['productId'] == productId:
                        chk_loop = "ok"
                        link_excel.cell(i, 8).value = itemCount
                        jisho_wb.save('./etc/jisho_link.xlsx')
                        break
                allCount += 1
            except:
                link_excel.cell(i, 8).value = "측정불가"
                jisho_wb.save('./etc/jisho_link.xlsx')
                break
    pg.alert(text="순위 검색이 완료 되었습니다.")


# 함수 시작염

# 상품 들어가서 스크롤 내리고 나오기

def onProductScroll(maxRange):
    pg.moveTo(200, 200)
    forCount = 0
    while maxRange > forCount:
        scrollVal = random.randrange(300, 500)
        pg.scroll(-scrollVal)
        wait_float(3, 5)
        forCount += 1

    while True:
        driver.back()
        wait_float(3, 5)
        get_shop_list = driver.find_element(by=By.CSS_SELECTOR, value='.mainFilter_option__c4_Lq')
        if get_shop_list is not None:
            return


# 지식쇼핑 검색 (2번 해야되니께~~)
def searchJisho(searchKeyword, driver):
    reCount = 0
    while True:
        reCount += 1
        wait_float(0.5,0.9)
        if reCount % 5 == 0:
            driver.refresh()
            pg.press('F5')
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='#sear')
            if(nShopSearchVar):
                break
        except:
            pass
        
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='#input_text')
            if(nShopSearchVar):
                break
        except:
            pass
        try:
            nShopSearchVar = driver.find_element(by=By.CSS_SELECTOR, value='._combineHeader_text_result_8IG-1')
            if(nShopSearchVar):
                break
        except:
            pass
        
    
    driver.execute_script("window.scrollTo(0,0);")
    wait_float(0.5, 1)
    nShopSearchVar.click()

    focus_window("네이버쇼핑")

    wait_float(0.5, 1)
    pg.hotkey('ctrl', 'a')
    wait_float(0.5, 1)
    pg.hotkey('del')
    wait_float(0.5, 1)

    # nShopSearchVar[0].send_keys(searchKeyword)
    # nShopSearchVar.click()
    keyboard.write(text=searchKeyword, delay=0.3)

    wait_float(1.2, 2.5)
    pg.hotkey('enter')
    wait_float(0.5, 1)
    pg.hotkey('enter')

    # 만약 에러나면 검색한 페이지에서 Element가 없어질때까지 loop 추가하자 (네이버 쇼핑 메인은 _lnb_infoscroll-view_1TdpI)

# 엑셀 내 목표클릭과 현재 클릭수 비교해서 배열에 값을 넣음


async def waitPrint(arr, ex, i, chk):
    target_click = int(ex.cell(i, 4).value)
    now_click = ex.cell(i, 5).value

    if now_click is None:
        ex.cell(i, 5).value = 0
        now_click = 0
    now_click = int(now_click)
    if chk == 'Y':
        if now_click < target_click:
            arr.append(i)
    else:
        if now_click >= target_click:
            arr.append(i)


async def playAsync_getArr(arr, ex, linkCount, chk):
    try:
        await asyncio.gather(*[waitPrint(arr, ex, i, chk) for i in range(1, linkCount + 1)])
    except:
        pass


# 결과 값(workarr) 을 가지고 해당 인덱스의 엑셀에 1씩 더하기
async def linkExcelPlus(ex, val):
    setVal = ex.cell(val, 5).value
    ex.cell(val, 5).value = setVal + 1


async def playAsync_plusArr(arr, ex):
    try:
        await asyncio.gather(*[linkExcelPlus(ex, val) for val in arr])
    except:
        pass


def naverLogin(load_id, load_pass):
    search_bar = searchElement(".sch_ico_aside")
    search_bar[0].click()
    login_btn = searchElement(".ss_profile_wrap")
    login_btn[0].click()

    # 로그인 부분

    focus_window('로그인')
    while True:
        searchElement("#id")

        pyperclip.copy(load_id)
        id_input = driver.find_elements(by=By.CSS_SELECTOR, value="#id")
        id_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)

        pyperclip.copy(load_pass)
        pw_input = driver.find_elements(by=By.CSS_SELECTOR, value="#pw")
        pw_input[0].click()
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'a')
        wait_float(0.4, 0.7)
        pg.hotkey('ctrl', 'v')
        wait_float(0.4, 0.7)
        id_input_value = id_input[0].get_attribute('value')
        if id_input_value:
            pg.hotkey('enter')
            wait_float(0.5, 1.0)
        else:
            continue

        asideChk = 0
        noProblem = ""
        passExit = ""
        while 2 > asideChk:
            try:
                waitAside = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".aside_wrap")))
                if waitAside is not None:
                    noProblem = "on"
                    break
            except:
                asideChk += 1

        if noProblem != "on":
            searchElement("#header")
        try:
            newDevice = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_white")
            newDevice[0].click()
        except:
            pass

        try:
            greenBtn = driver.find_elements(
                by=By.CSS_SELECTOR, value=".btn_next")
            greenBtn[0].click()
            passExit = "on"
        except:
            pass

        try:
            protectId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".ico_warning2")
            if protectId:
                return "보호조치"
        except:
            pass

        try:
            sleepId = driver.find_elements(
                by=By.CSS_SELECTOR, value=".warning_v2")
            if sleepId:
                return "휴면아이디"
        except:
            pass

        try:
            unPwd = driver.find_elements(
                by=By.CSS_SELECTOR, value=".error_message")
            if unPwd:
                return "비번틀림"
            # 다시 로그인 어쩌구......
        except:
            pass
        
        try:
            unPwd = driver.find_elements(by=By.CSS_SELECTOR, value=".action_inner")
            if unPwd:
                return "비정상적 활동"
            # 다시 로그인 어쩌구......
        except:
            pass

        if passExit != "on":
            goToMain = searchElement(".ah_close")
            goToMain[0].click()
            wait_float(0.4, 0.7)

        time.sleep(5)
        break
        # 로그인 부분 끝


def changeIp():
    try:
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        ondevice = device[0]
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except:

        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("http://ip.jsontest.com").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp


def changeIpSpeed():
    os.system('adb server start')
    client = AdbClient(host="127.0.0.1", port=5037)
    device = client.devices()  # 디바이스 1개
    ondevice = device[0]
    while True:
        try:
            
            ondevice.shell("input keyevent KEYCODE_POWER")
            ondevice.shell("svc data disable")
            ondevice.shell("settings put global airplane_mode_on 1")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

            ondevice.shell("svc data enable")
            ondevice.shell("settings put global airplane_mode_on 0")
            ondevice.shell(
                "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
            time.sleep(3)
            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
        except:

            while True:
                try:
                    wait_float(0.5, 0.9)
                    getIp = requests.get("http://ip.jsontest.com").json()['ip']
                    if getIp is not None:
                        break
                except:
                    continue
                
        
        driver.get('https://fast.com/ko/')
        searchElement('.speed-results-container')
        time.sleep(3)
        getInternetRapidEle = searchElement('.speed-results-container')
        getInternetRapid = getInternetRapidEle[0].text
        if float(getInternetRapid) < 2.7:
            continue
        else:
            driver.close()
            break
        
    return getIp

# def searchElement(ele):
#     time.sleep(1)
#     re_count = 1
#     element = ""
#     while True:
#         if re_count % 5 == 0:
#             print(ele)
#             print("새로고침!!!!")
#             driver.refresh()
#         elif element != "":
#             break
#         try:
#             element = WebDriverWait(driver, 1).until(
#                 EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
#         except:
#             re_count += 1

#     selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
#     wait_float(0.5, 1.2)
#     return selected_element

def getUaNum():
    with open("./etc/useragent/useragent_all.txt", "r") as f:
        fArr = f.readlines()
        fCount = len(fArr)
        uaSet = random.randrange(0, fCount)
    return uaSet

def searchElement(ele):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            focus_window('chrome')
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element


def untilEleShow(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        try:
            btnEle = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is not None:
                return
        except:
            continue


def untilEleGone(clickEle, searchEle):
    while True:
        try:
            clickEle.click()
            time.sleep(1)
        except:
            pass
        
        try:
            btnEle = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, searchEle)))
            if btnEle is None:
                return
        except:
            return


def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp():
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    if winName == 'chkname':
        win_list = gw.getAllTitles()
        # pg.alert(text=f"{win_list}")
    # 윈도우 타이틀에 Chrome 이 포함된 모든 윈도우 수집, 리스트로 리턴
    win = gw.getWindowsWithTitle(winName)[0]
    win.activate()  # 윈도우 활성화


BASE_DIR = Path(__file__).resolve().parent


def get_secret(
    key: str,
    default_value: Optional[str] = None,
    json_path: str = str(BASE_DIR / "secrets.json"),
):

    with open(json_path) as f:
        secrets = json.loads(f.read())
    try:
        return secrets[key]
    except KeyError:
        if default_value:
            return default_value
        raise EnvironmentError(f"Set the {key} environment variable.")
